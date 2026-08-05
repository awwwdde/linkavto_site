from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from urllib.parse import quote, urlencode
import random
import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from shop.models import Product, Category, ProductClick, ProductView, Review, CarBrand, TruckBrand, MotoBrand, SpecialBrand, CarModel, TruckModel, MotoModel, SpecialModel, CarGeneration, TruckGeneration, MotoGeneration, SpecialGeneration, CarModification, TruckModification, MotoModification, SpecialModification, Manufacturer, CategoryRequest
from .models import Seller, SellerCategory, SellerReviewSummary, ProductReview, ReviewReport, \
    SellerNotification, SellerIntegration, SellerSettings
from django.core.paginator import Paginator
import csv
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import (
    Count, Sum, Avg, F, Q, ExpressionWrapper,
    DecimalField, DateTimeField, DateField, Count
)
from django.utils import timezone
from .models import FinancialReport, Transaction, Decimal, Payout
from orders.models import Order, OrderItem
from favorites.models import Favorite
from accounts.models import Profile
from django.db.models import Sum, Count, Q
from datetime import datetime, timedelta
from django.views.generic import TemplateView, ListView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Count, Avg
from django.shortcuts import get_object_or_404
from django.views.generic import TemplateView
from django.http import JsonResponse
from django.urls import reverse
import json
from .advanced_analytics import (
    SalesPredictor, CustomerLTV, AnomalyDetector,
    DynamicPricing, ProductSegmentation, MarketingAnalyzer,
    SmartAlertSystem, BehaviorAnalyzer
)
from django.db.models.functions import TruncMonth, TruncWeek
from .models import SellerFinanceProfile
from garage.services import VINService

logger = logging.getLogger(__name__)


def _get_validated_product_vin(data):
    vin = VINService.normalize_vin(data.get('vin_number', ''))
    if not vin:
        return None, None

    vin_validation = VINService.validate_vin(vin)
    if not vin_validation['valid']:
        return None, vin_validation['error']

    return vin_validation.get('normalized_vin') or vin, None

def seller_storefront(request, seller_id: int):
    """
    Публичная витрина продавца для покупателей.
    Не требует прав продавца и НЕ ведёт в кабинет (/sellers/products/).
    """
    seller = get_object_or_404(
        Seller.objects.select_related('review_summary'),
        id=seller_id
    )

    def _get_multi_param(key: str) -> list:
        raw_list = request.GET.getlist(key) or []
        if not raw_list:
            raw = request.GET.get(key)
            raw_list = [raw] if raw else []
        out: list = []
        for item in raw_list:
            if not item:
                continue
            for part in str(item).split(','):
                val = part.strip()
                if val and val not in out:
                    out.append(val)
        return out

    products_qs = (
        Product.objects
        .filter(seller=seller, is_active=True)
        .select_related('category', 'manufacturer', 'seller', 'car_brand', 'truck_brand', 'moto_brand', 'special_brand')
    )

    # Поиск по названию, артикулу, VIN
    search_query = request.GET.get('search', '').strip()
    if search_query:
        products_qs = products_qs.filter(
            Q(name__icontains=search_query)
            | Q(part_number__icontains=search_query)
            | Q(vin__icontains=search_query)
        )

    # --- Фильтр по разделу (show_in) и категории ---
    from shop.models import Category as ShopCategory
    from django.db.models import Q as _Q

    section_slug = request.GET.get('section', '').strip()
    category_slug = request.GET.get('category', '').strip()
    part_subcategory_slug = request.GET.get('part_subcategory', '').strip()
    part_subsubcategory_slug = request.GET.get('part_subsubcategory', '').strip()

    # Доступные разделы у данного продавца
    available_show_in = (
        ShopCategory.objects
        .filter(products__seller=seller, products__is_active=True)
        .order_by()
        .values_list('show_in', flat=True)
        .distinct()
    )
    SHOW_IN_MAP = dict(ShopCategory.SHOW_IN_CHOICES)
    available_sections = [
        {'code': code, 'name': SHOW_IN_MAP[code]}
        for code in sorted(available_show_in)
        if code in SHOW_IN_MAP
    ]

    if section_slug:
        products_qs = products_qs.filter(category__show_in=section_slug)

    seller_categories = (
        ShopCategory.objects
        .filter(products__seller=seller, products__is_active=True)
        .distinct()
        .order_by('name')
    )
    if section_slug:
        seller_categories = seller_categories.filter(show_in=section_slug)

    # Deduplicate by name — same category name can appear across different show_in sections
    _seen_cat_names = set()
    _unique_cats = []
    for _cat in seller_categories:
        if _cat.name not in _seen_cat_names:
            _seen_cat_names.add(_cat.name)
            _unique_cats.append(_cat)
    seller_categories = _unique_cats

    # Все подкатегории для всех категорий продавца (для клиентской фильтрации)
    _seller_cat_slugs = [c.slug for c in seller_categories]
    all_seller_subcategories = list(
        ShopCategory.objects.filter(
            parent__slug__in=_seller_cat_slugs,
            is_active=True
        ).select_related('parent').order_by('name')
    ) if _seller_cat_slugs else []

    if category_slug:
        products_qs = products_qs.filter(category__slug=category_slug)

    # Подкатегории / под-подкатегории
    if part_subsubcategory_slug:
        products_qs = products_qs.filter(category__slug=part_subsubcategory_slug)
    elif part_subcategory_slug:
        products_qs = products_qs.filter(
            _Q(category__slug=part_subcategory_slug) |
            _Q(category__parent__slug=part_subcategory_slug)
        )

    # Строим списки для фильтра
    subcategories = []
    sub_subcategories = []
    if category_slug:
        try:
            _selected_cat = ShopCategory.objects.get(slug=category_slug, is_active=True)
            subcategories = list(_selected_cat.children.filter(is_active=True).order_by('name'))
        except ShopCategory.DoesNotExist:
            pass
    if part_subcategory_slug and subcategories:
        try:
            _selected_subcat = ShopCategory.objects.get(slug=part_subcategory_slug, is_active=True)
            sub_subcategories = list(_selected_subcat.children.filter(is_active=True).order_by('name'))
        except ShopCategory.DoesNotExist:
            pass


    # --- Фильтры ---
    brand_slugs = _get_multi_param('brand')
    model_slugs = _get_multi_param('model')
    generation_slugs = _get_multi_param('generation')
    manufacturer_slugs = _get_multi_param('manufacturer')
    price_min = request.GET.get('price_min')
    price_max = request.GET.get('price_max')
    in_stock = request.GET.get('in_stock')
    on_order = request.GET.get('on_order')
    is_original = request.GET.get('is_original')

    need_distinct = False

    if brand_slugs:
        brand_q = Q()
        for b in brand_slugs:
            brand_q |= (
                Q(car_brand__slug=b) |
                Q(truck_brand__slug=b) |
                Q(moto_brand__slug=b) |
                Q(special_brand__slug=b)
            )
        products_qs = products_qs.filter(brand_q)
        need_distinct = True

    if model_slugs:
        model_q = (
            Q(car_models__slug__in=model_slugs) |
            Q(truck_models__slug__in=model_slugs) |
            Q(moto_models__slug__in=model_slugs) |
            Q(special_models__slug__in=model_slugs)
        )
        products_qs = products_qs.filter(model_q)
        need_distinct = True

    if generation_slugs:
        gen_q = (
            Q(car_generations__slug__in=generation_slugs) |
            Q(truck_generations__slug__in=generation_slugs) |
            Q(moto_generations__slug__in=generation_slugs) |
            Q(special_generations__slug__in=generation_slugs)
        )
        products_qs = products_qs.filter(gen_q)
        need_distinct = True

    if manufacturer_slugs:
        products_qs = products_qs.filter(manufacturer__slug__in=manufacturer_slugs)

    if price_min:
        try:
            products_qs = products_qs.filter(price__gte=float(price_min))
        except (ValueError, TypeError):
            pass

    if price_max:
        try:
            products_qs = products_qs.filter(price__lte=float(price_max))
        except (ValueError, TypeError):
            pass

    if in_stock and not on_order:
        products_qs = products_qs.filter(stock__gt=0)
    elif on_order and not in_stock:
        products_qs = products_qs.filter(stock=0)

    if is_original:
        if str(is_original).lower() not in ('false', '0', 'no', 'off', ''):
            products_qs = products_qs.filter(is_original=True)

    if need_distinct:
        products_qs = products_qs.distinct()

    # Сортировка
    sort_map = {
        'price_asc': 'price',
        'price_desc': '-price',
        'newest': '-created_at',
        'popular': '-created_at',
    }
    sort_key = request.GET.get('sort', 'newest')
    order_field = sort_map.get(sort_key, '-created_at')
    products_qs = products_qs.order_by(order_field)

    paginator = Paginator(products_qs, 24)
    page_obj = paginator.get_page(request.GET.get('page'))

    favorite_ids = []
    if request.user.is_authenticated:
        favorite_ids = list(
            Favorite.objects.filter(user=request.user).values_list('product_id', flat=True)
        )

    product_count = Product.objects.filter(seller=seller, is_active=True).count()

    # --- Данные для фильтров ---
    # Подбираем бренды по разделу (как в основном каталоге), фильтруем по товарам продавца
    if section_slug == 'trucks':
        _seller_brand_ids = (
            Product.objects
            .filter(seller=seller, is_active=True, truck_brand__isnull=False)
            .values_list('truck_brand_id', flat=True)
            .distinct()
        )
        brands = TruckBrand.objects.filter(id__in=_seller_brand_ids, is_active=True).order_by('name')
    elif section_slug == 'moto':
        _seller_brand_ids = (
            Product.objects
            .filter(seller=seller, is_active=True, moto_brand__isnull=False)
            .values_list('moto_brand_id', flat=True)
            .distinct()
        )
        brands = MotoBrand.objects.filter(id__in=_seller_brand_ids, is_active=True).order_by('name')
    elif section_slug == 'special':
        _seller_brand_ids = (
            Product.objects
            .filter(seller=seller, is_active=True, special_brand__isnull=False)
            .values_list('special_brand_id', flat=True)
            .distinct()
        )
        brands = SpecialBrand.objects.filter(id__in=_seller_brand_ids, is_active=True).order_by('name')
    else:
        # cars или любой другой / не выбран
        _seller_brand_ids = (
            Product.objects
            .filter(seller=seller, is_active=True, car_brand__isnull=False)
            .values_list('car_brand_id', flat=True)
            .distinct()
        )
        brands = CarBrand.objects.filter(id__in=_seller_brand_ids, is_active=True).order_by('name')

    # Загружаем все модели/поколения для клиентской фильтрации (как в category_view)
    # Фильтруем только по брендам продавца, чтобы не грузить весь справочник
    _brand_ids = brands.values_list('id', flat=True)
    if section_slug == 'trucks':
        models_qs = TruckModel.objects.filter(brand_id__in=_brand_ids, is_active=True).select_related('brand').order_by('brand__name', 'name')
        gens_qs = TruckGeneration.objects.filter(model__brand_id__in=_brand_ids, is_active=True).select_related('model').order_by('model__name', '-year_start')
    elif section_slug == 'moto':
        models_qs = MotoModel.objects.filter(brand_id__in=_brand_ids, is_active=True).select_related('brand').order_by('brand__name', 'name')
        gens_qs = MotoGeneration.objects.filter(model__brand_id__in=_brand_ids, is_active=True).select_related('model').order_by('model__name', '-year_start')
    elif section_slug == 'special':
        models_qs = SpecialModel.objects.filter(brand_id__in=_brand_ids, is_active=True).select_related('brand').order_by('brand__name', 'name')
        gens_qs = SpecialGeneration.objects.filter(model__brand_id__in=_brand_ids, is_active=True).select_related('model').order_by('model__name', '-year_start')
    else:
        models_qs = CarModel.objects.filter(brand_id__in=_brand_ids, is_active=True).select_related('brand').order_by('brand__name', 'name')
        gens_qs = CarGeneration.objects.filter(model__brand_id__in=_brand_ids, is_active=True).select_related('model').order_by('model__name', '-year_start')

    # Список модификаций — загружаем все для брендов продавца (фильтрация на стороне клиента)
    if section_slug == 'trucks':
        mods_qs = TruckModification.objects.filter(
            generation__model__brand_id__in=_brand_ids
        ).select_related('generation').order_by('generation__name', 'name')
    elif section_slug == 'moto':
        mods_qs = MotoModification.objects.filter(
            generation__model__brand_id__in=_brand_ids
        ).select_related('generation').order_by('generation__name', 'name')
    elif section_slug == 'special':
        mods_qs = SpecialModification.objects.filter(
            generation__model__brand_id__in=_brand_ids
        ).select_related('generation').order_by('generation__name', 'name')
    else:
        mods_qs = CarModification.objects.filter(
            generation__model__brand_id__in=_brand_ids
        ).select_related('generation').order_by('generation__name', 'name')

    manufacturer_ids = (
        Product.objects
        .filter(seller=seller, is_active=True)
        .exclude(manufacturer_id__isnull=True)
        .values_list('manufacturer_id', flat=True)
        .distinct()
    )
    manufacturers = Manufacturer.objects.filter(is_active=True, id__in=manufacturer_ids).order_by('name')

    return render(request, 'sellers/seller_storefront.html', {
        'seller': seller,
        'products': page_obj,
        'favorite_ids': favorite_ids,
        'product_count': product_count,
        'search_query': search_query,
        'current_sort': sort_key,
        'brands': brands,
        'models': models_qs,
        'generations': gens_qs,
        'modifications': mods_qs,
        'manufacturers': manufacturers,
        'current_vehicle_type': 'car',
        'subcategories': subcategories,
        'sub_subcategories': sub_subcategories,
        'seller_categories': seller_categories,
        'all_seller_subcategories': all_seller_subcategories,
        'current_category_slug': category_slug,
        'available_sections': available_sections,
        'current_section': section_slug,
        'current_part_subcategory': part_subcategory_slug,
        'current_part_subsubcategory': part_subsubcategory_slug,
    })


def become_seller(request):
    return render(request, 'sellers/become_seller.html')


def seller_login(request):
    return render(request, 'sellers/seller_login.html')


@login_required
def seller_dashboard(request):
    """Главная страница кабинета продавца"""
    # Проверяем, есть ли у пользователя профиль продавца
    if not hasattr(request.user, 'seller'):
        messages.warning(request, 'У вас нет профиля продавца. Пожалуйста, зарегистрируйтесь как продавец.')
        return redirect('sellers:seller_registration')
    
    seller = request.user.seller

    # Объявляем переменные заранее
    today = timezone.now().date()
    finance_profile = None
    today_sales = 0
    today_orders = 0
    active_products = 0
    avg_rating = 4.8
    total_reviews = 0
    pending_payouts = 0
    month_sales = 0

    # Получаем финансовые данные для карточек
    try:
        finance_profile = SellerFinanceProfile.objects.get(seller=seller)

        # Выручка за сегодня
        today_sales = Transaction.objects.filter(
            seller=seller,
            transaction_type='sale',
            transaction_date__date=today,
            status='completed'
        ).aggregate(total_sales=Sum('net_amount'))['total_sales'] or 0

        # Новые заказы (за сегодня)
        today_orders = Order.objects.filter(
            items__product__seller=seller,
            created_at__date=today
        ).distinct().count()

        # Товары в продаже (активные и в наличии)
        active_products = Product.objects.filter(
            seller=seller,
            is_active=True,
            stock__gt=0
        ).count()

        # Рейтинг магазина из отзывов
        try:
            review_summary = SellerReviewSummary.objects.get(seller=seller)
            avg_rating = float(review_summary.average_rating) if review_summary.average_rating else 4.8
            total_reviews = review_summary.total_reviews
        except SellerReviewSummary.DoesNotExist:
            pass  # Используем значения по умолчанию

        # Ожидающие выплаты
        pending_payouts = Payout.objects.filter(
            seller=seller,
            status__in=['pending', 'processing']
        ).aggregate(total_payouts=Sum('amount'))['total_payouts'] or 0

        # Продажи за 30 дней
        month_ago = today - timedelta(days=30)
        month_sales = Transaction.objects.filter(
            seller=seller,
            transaction_type='sale',
            transaction_date__date__gte=month_ago,
            status='completed'
        ).aggregate(total_sales=Sum('net_amount'))['total_sales'] or 0

    except SellerFinanceProfile.DoesNotExist:
        # Используем значения по умолчанию, которые уже установлены выше
        pass

    # Расчет изменений для заказов
    yesterday = today - timedelta(days=1)
    yesterday_orders = Order.objects.filter(
        items__product__seller=seller,
        created_at__date=yesterday
    ).distinct().count()

    orders_change = 0
    if yesterday_orders > 0:
        orders_change = ((today_orders - yesterday_orders) / yesterday_orders) * 100

    # Расчет изменений для товаров
    week_ago = today - timedelta(days=7)
    products_week_ago = Product.objects.filter(
        seller=seller,
        is_active=True,
        stock__gt=0,
        created_at__date__lte=week_ago
    ).count()

    products_change = 0
    if products_week_ago > 0:
        products_change = ((active_products - products_week_ago) / products_week_ago) * 100

    # Расчет изменений для рейтинга
    rating_change = get_rating_change(seller, avg_rating)

    # Получаем последние заказы для dashboard (5 последних)
    recent_orders = Order.objects.filter(
        items__product__seller=seller
    ).distinct().prefetch_related(
        'items',
        'items__product'
    ).annotate(
        seller_items_count=Count('items', filter=Q(items__product__seller=seller)),
        seller_total=Sum(
            F('items__price') * F('items__quantity'),
            filter=Q(items__product__seller=seller)
        )
    ).order_by('-created_at')[:5]

    context = {
        'finance_profile': finance_profile,
        'today_sales': today_sales,
        'today_orders': today_orders,
        'orders_change': orders_change,
        'active_products': active_products,
        'products_change': products_change,
        'avg_rating': round(avg_rating, 1),
        'rating_change': rating_change,
        'total_reviews': total_reviews,
        'pending_payouts': pending_payouts,
        'month_sales': month_sales,
        'recent_orders': recent_orders,  # Добавляем последние заказы
    }

    return render(request, 'sellers/seller_dashboard.html', context)


@csrf_exempt
def send_verification_code(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email', '').strip()

            if not email:
                return JsonResponse({'success': False, 'error': 'Email обязателен'})

            # Проверяем, что это email
            if '@' not in email:
                return JsonResponse({'success': False, 'error': 'Введите корректный email'})

            # Проверяем, не отправлялся ли код недавно (в течение 30 секунд)
            from django.utils import timezone
            from datetime import timedelta
            from accounts.models import VerificationCode

            recent_codes = VerificationCode.objects.filter(
                login=email,
                created_at__gte=timezone.now() - timedelta(seconds=30)
            )

            if recent_codes.exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Код уже отправлен. Подождите 30 секунд перед повторной отправкой.'
                }, status=429)

            # Генерируем 6-значный код
            code = ''.join([str(random.randint(0, 9)) for _ in range(6)])

            # Сохраняем код в базе данных
            try:
                VerificationCode.objects.create(
                    login=email,
                    code=code
                )
                logger.info(f"Verification code saved for {email}")
            except Exception as e:
                logger.error(f"Error saving verification code: {e}")
                return JsonResponse({'success': False, 'error': 'Ошибка сохранения кода'}, status=500)

            # Отправляем email
            email_sent = False
            try:
                site_url = getattr(settings, 'SITE_URL', 'https://linkavto.ru')
                image_url = f"{site_url}/static/img/{quote('Баннер легковые авто 1920x1080.png')}"
                verify_url = f"{site_url}/account/verify-by-link/?{urlencode({'code': code, 'login': email})}"
                html_message = render_to_string('accounts/emails/verification_code.html', {
                    'code': code,
                    'site_url': site_url,
                    'image_url': image_url,
                    'button_url': verify_url,
                })
                plain_message = f'''Здравствуйте!

Ваш код подтверждения: {code}

Код действителен в течение 10 минут.
Если вы не запрашивали код подтверждения, проигнорируйте это письмо.
С уважением,
LINKAVTO'''
                send_mail(
                    'Код подтверждения для регистрации продавца в LinkAvto',
                    plain_message,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                    html_message=html_message,
                )
                email_sent = True
                logger.info(f"Verification code email sent successfully to {email}")
            except Exception as e:
                logger.warning(f"Failed to send email to {email}: {e}")
                # В режиме разработки показываем код в консоли
                print(f"DEV: Verification code for {email}: {code}")
                logger.info(f"Verification code {code} saved for {email} (email not sent, code shown in console)")

            # Возвращаем успешный ответ даже если email не отправился (для режима разработки)
            # Код сохранен в базе и может быть использован для проверки
            response_data = {
                'success': True,
                'message': 'Код отправлен на email' if email_sent else 'Код сохранен (email не отправлен)'
            }
            
            # В режиме разработки можно добавить код в ответ (опционально, для тестирования)
            if not email_sent and settings.DEBUG:
                response_data['debug_code'] = code
            
            return JsonResponse(response_data)

        except Exception as e:
            logger.error(f"Error sending verification code: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Ошибка сервера'})

    return JsonResponse({'success': False, 'error': 'Метод не разрешен'})


@csrf_exempt
def verify_code(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email', '').strip()
            code = data.get('code', '').strip()

            if not email or not code:
                return JsonResponse({'success': False, 'error': 'Email и код обязательны'})

            from django.utils import timezone
            from accounts.models import VerificationCode

            # Ищем актуальный код
            try:
                verification_code = VerificationCode.objects.filter(
                    login=email,
                    is_used=False,
                    created_at__gte=timezone.now() - timezone.timedelta(minutes=10)
                ).latest('created_at')

                logger.info(f"Found verification code for {email}")

            except VerificationCode.DoesNotExist:
                logger.warning(f"Verification code not found or expired for {email}")
                return JsonResponse({'success': False, 'error': 'Код не найден или устарел'}, status=400)

            # Проверяем код
            if verification_code.code != code:
                verification_code.attempts -= 1
                verification_code.save()

                if verification_code.attempts <= 0:
                    verification_code.is_used = True
                    verification_code.save()
                    return JsonResponse({'success': False, 'error': 'Превышено количество попыток'}, status=400)

                return JsonResponse({
                    'success': False,
                    'error': 'Неверный код',
                    'attempts_left': verification_code.attempts
                }, status=400)

            # Код верный, помечаем как использованный
            verification_code.is_used = True
            verification_code.save()
            logger.info(f"Code verified successfully for {email}")

            # Сохраняем подтвержденный email в сессии
            request.session['registration_email'] = email
            request.session['email_verified'] = True

            return JsonResponse({'success': True})

        except Exception as e:
            logger.error(f"Error verifying code: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Ошибка сервера'})

    return JsonResponse({'success': False, 'error': 'Метод не разрешен'})


@csrf_exempt
def seller_login_with_code(request):
    """API для входа продавца по коду подтверждения"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email', '').strip()
            code = data.get('code', '').strip()

            if not email or not code:
                return JsonResponse({'success': False, 'error': 'Email и код обязательны'})

            from django.utils import timezone
            from accounts.models import VerificationCode
            from django.contrib.auth import login

            # Ищем актуальный код
            try:
                verification_code = VerificationCode.objects.filter(
                    login=email,
                    is_used=False,
                    created_at__gte=timezone.now() - timezone.timedelta(minutes=10)
                ).latest('created_at')

                logger.info(f"Found verification code for seller login: {email}")

            except VerificationCode.DoesNotExist:
                logger.warning(f"Verification code not found or expired for seller login: {email}")
                return JsonResponse({'success': False, 'error': 'Код не найден или устарел'}, status=400)

            # Проверяем код
            if verification_code.code != code:
                verification_code.attempts -= 1
                verification_code.save()

                if verification_code.attempts <= 0:
                    verification_code.is_used = True
                    verification_code.save()
                    return JsonResponse({'success': False, 'error': 'Превышено количество попыток'}, status=400)

                return JsonResponse({
                    'success': False,
                    'error': 'Неверный код',
                    'attempts_left': verification_code.attempts
                }, status=400)

            # Код верный, помечаем как использованный
            verification_code.is_used = True
            verification_code.save()
            logger.info(f"Code verified successfully for seller login: {email}")

            # Ищем пользователя по email
            try:
                user = User.objects.get(email=email)
            except User.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'error': 'Продавец с таким email не найден. Пожалуйста, зарегистрируйтесь.'
                }, status=400)

            # Проверяем, есть ли у пользователя профиль продавца
            try:
                seller = user.seller
            except Seller.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'error': 'У вас нет профиля продавца. Пожалуйста, зарегистрируйтесь как продавец.'
                }, status=400)

            # Авторизуем пользователя
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            logger.info(f"Seller logged in successfully: {email}")

            return JsonResponse({
                'success': True,
                'redirect_url': '/sellers/seller-dashboard/'
            })

        except Exception as e:
            logger.error(f"Error in seller login: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Ошибка сервера'})

    return JsonResponse({'success': False, 'error': 'Метод не разрешен'})



def seller_registration(request):
    if request.method == 'POST':
        step = request.POST.get('step')

        if step == '0':  # Шаг email
            email = request.POST.get('email')
            if email:
                # Проверяем, нет ли уже пользователя с таким email
                if User.objects.filter(email=email).exists():
                    return JsonResponse({
                        'success': False,
                        'error': 'Пользователь с таким email уже существует'
                    })

                request.session['registration_email'] = email
                request.session['registration_step'] = 0
                return JsonResponse({'success': True})
            return JsonResponse({'success': False, 'error': 'Email обязателен'})

        elif step == '1':  # Шаг основной информации
            email = request.session.get('registration_email')
            email_verified = request.session.get('email_verified', False)
            
            if not email:
                return JsonResponse({'success': False, 'error': 'Сначала введите email'})
            
            if not email_verified:
                return JsonResponse({'success': False, 'error': 'Сначала подтвердите email'})

            # Валидация полей шага 1
            required_fields = {
                'contact_person': 'ФИО контактного лица',
                'phone': 'Телефон',
                'company_name': 'Название компании',
                'store_name': 'Название магазина',
                'legal_form': 'Форма собственности',
                'inn': 'ИНН',
                'ogrn': 'ОГРН/ОГРНИП',
                'legal_address': 'Юридический адрес',
                'shipping_address': 'Адрес отправки'
            }

            errors = {}
            phone = request.POST.get('phone')

            # Проверяем обязательные поля
            for field, field_name in required_fields.items():
                value = request.POST.get(field)
                if not value:
                    errors[field] = f'{field_name} обязателен для заполнения'

            # Проверяем уникальность ИНН
            inn = request.POST.get('inn')
            if inn and Seller.objects.filter(inn=inn).exists():
                errors['inn'] = 'Продавец с таким ИНН уже зарегистрирован'

            if errors:
                return JsonResponse({
                    'success': False,
                    'errors': errors
                })

            try:
                # ВРЕМЕННО убираем transaction.atomic для диагностики
                # with transaction.atomic():

                # Получаем полные данные из DaData API для сохранения в api_data
                api_data_text = None
                if inn:
                    try:
                        from sellers.api_services.dadata_service import DaDataService
                        dadata_service = DaDataService()
                        full_company_data = dadata_service.get_company_by_inn(inn)
                        
                        if full_company_data and full_company_data.get('raw_api_data'):
                            # Форматируем данные в текстовый формат "key: val\nkey: val"
                            api_data_text = format_dadata_api_data(full_company_data.get('raw_api_data'))
                    except Exception as api_error:
                        logger.warning(f"Не удалось получить данные из DaData API для сохранения: {api_error}")

                # Создаем или получаем пользователя
                user, user_created = User.objects.get_or_create(
                    email=email,
                    defaults={
                        'username': email,
                        'is_active': True,
                        'first_name': request.POST.get('contact_person', '').split(' ')[0] if request.POST.get(
                            'contact_person') else '',
                    }
                )

                # Если пользователь уже существовал, проверяем не привязан ли он к другому продавцу
                if not user_created and hasattr(user, 'seller'):
                    return JsonResponse({
                        'success': False,
                        'error': 'Этот email уже привязан к другому продавцу'
                    })

                # Если пользователь создан, устанавливаем пароль по умолчанию
                if user_created:
                    # Генерируем случайный пароль
                    import secrets
                    import string
                    alphabet = string.ascii_letters + string.digits
                    password = ''.join(secrets.choice(alphabet) for i in range(12))
                    user.set_password(password)
                    user.save()
                    print(f"Создан пользователь: {email}")

                # ВАЛИДАЦИЯ ДАННЫХ ПРОДАВЦА ПЕРЕД СОЗДАНИЕМ
                seller_data = {
                    'user': user,
                    'contact_person': request.POST.get('contact_person'),
                    'phone': phone,
                    'company_name': request.POST.get('company_name'),
                    'store_name': request.POST.get('store_name', '').strip() or request.POST.get('company_name', ''),
                    'legal_form': request.POST.get('legal_form'),
                    'inn': inn,
                    'ogrn': request.POST.get('ogrn'),
                    'legal_address': request.POST.get('legal_address'),
                    'actual_address': request.POST.get('actual_address', ''),
                    'shipping_address': request.POST.get('shipping_address'),
                    'website_url': request.POST.get('website_url', ''),
                    'product_info': request.POST.get('product_info', ''),
                    'status': 'pending',
                    'is_active': True
                }
                
                # Сохраняем данные из API ФНС в текстовом формате
                if api_data_text:
                    seller_data['api_data'] = {'text': api_data_text}

                # Проверяем данные вручную
                print("Данные для создания продавца:")
                for key, value in seller_data.items():
                    if key != 'user':  # Не выводим объект user
                        print(f"  {key}: {value}")

                # Создаем продавца
                seller = Seller(**seller_data)

                # Пробуем сохранить без full_clean сначала
                try:
                    seller.save()
                    print("Продавец успешно создан")
                except Exception as save_error:
                    print(f"Ошибка при сохранении продавца: {save_error}")
                    print(f"Тип ошибки: {type(save_error)}")

                    # Пробуем full_clean для получения конкретных ошибок валидации
                    try:
                        seller.full_clean()
                    except Exception as validation_error:
                        print(f"Ошибки валидации: {validation_error}")
                        return JsonResponse({
                            'success': False,
                            'error': f'Ошибки в данных: {validation_error}'
                        })

                    raise save_error

                # Сохраняем ID продавца в сессии
                request.session['seller_id'] = seller.id
                request.session['registration_step'] = 1
                
                # Авторизуем пользователя после успешной регистрации
                from django.contrib.auth import login
                login(request, user)

                return JsonResponse({'success': True, 'seller_id': seller.id})

            except Exception as e:
                print(f"Общая ошибка создания продавца: {e}")
                print(f"Тип ошибки: {type(e)}")

                # Более детальный анализ ошибки
                error_message = str(e)
                if 'UNIQUE' in error_message:
                    if 'phone' in error_message:
                        return JsonResponse({
                            'success': False,
                            'errors': {'phone': 'Продавец с таким телефоном уже зарегистрирован'}
                        })
                    elif 'inn' in error_message:
                        return JsonResponse({
                            'success': False,
                            'errors': {'inn': 'Продавец с таким ИНН уже зарегистрирован'}
                        })
                    elif 'user' in error_message:
                        return JsonResponse({
                            'success': False,
                            'error': 'Этот пользователь уже зарегистрирован как продавец'
                        })

                return JsonResponse({
                    'success': False,
                    'error': f'Ошибка сохранения: {error_message}'
                })


        elif step == '2':  # Шаг категорий

            seller_id = request.session.get('seller_id')

            if not seller_id:
                return JsonResponse({

                    'success': False,

                    'error': 'Сессия истекла. Пожалуйста, начните регистрацию заново.'

                })

            try:

                seller = Seller.objects.get(id=seller_id)

                categories = request.POST.getlist('categories')

                product_count = request.POST.get('product_count')

                # Валидация категорий

                if not categories:
                    return JsonResponse({

                        'success': False,

                        'errors': {'categories': 'Выберите хотя бы одну категорию'}

                    })

                # Валидация количества товаров

                if not product_count:
                    return JsonResponse({

                        'success': False,

                        'errors': {'product_count': 'Укажите количество товаров'}

                    })

                # Проверяем существование категорий в shop

                valid_categories = []

                for category_id in categories:

                    try:

                        category = Category.objects.get(id=category_id, is_active=True)  # Теперь из shop

                        valid_categories.append(category)

                    except Category.DoesNotExist:

                        return JsonResponse({

                            'success': False,

                            'error': f'Категория с ID {category_id} не найдена'

                        })

                # Сохраняем категории

                for category in valid_categories:
                    SellerCategory.objects.get_or_create(seller=seller, category=category)

                # Сохраняем количество товаров

                seller.product_count = product_count

                seller.save()

                # Очищаем сессию

                session_keys = ['registration_email', 'seller_id', 'registration_step']

                for key in session_keys:

                    if key in request.session:
                        del request.session[key]

                return JsonResponse({'success': True})


            except Seller.DoesNotExist:

                return JsonResponse({'success': False, 'error': 'Продавец не найден'})

            except Exception as e:

                print(f"Error saving categories: {e}")

                return JsonResponse({'success': False, 'error': f'Ошибка сохранения категорий: {str(e)}'})

    # GET запрос - отображаем форму
    current_step = request.session.get('registration_step', 0)
    context = {'current_step': current_step}
    return render(request, 'sellers/seller_registration.html', context)


def format_dadata_api_data(raw_data):
    """
    Форматирует данные из DaData API в текстовый формат "key: val\nkey: val"
    
    Args:
        raw_data: Словарь с данными из DaData API
        
    Returns:
        Строка с отформатированными данными
    """
    if not raw_data:
        return ""
    
    lines = []
    
    def format_value(value):
        """Рекурсивно форматирует значение"""
        if isinstance(value, dict):
            return format_dict(value)
        elif isinstance(value, list):
            return format_list(value)
        elif value is None:
            return ""
        else:
            return str(value)
    
    def format_dict(d, prefix=""):
        """Форматирует словарь"""
        for key, value in d.items():
            if isinstance(value, dict):
                lines.append(f"{prefix}{key}:")
                format_dict(value, prefix + "  ")
            elif isinstance(value, list):
                lines.append(f"{prefix}{key}:")
                format_list(value, prefix + "  ")
            else:
                val_str = format_value(value)
                if val_str:
                    lines.append(f"{prefix}{key}: {val_str}")
    
    def format_list(lst, prefix=""):
        """Форматирует список"""
        for i, item in enumerate(lst):
            if isinstance(item, dict):
                lines.append(f"{prefix}[{i}]:")
                format_dict(item, prefix + "  ")
            else:
                val_str = format_value(item)
                if val_str:
                    lines.append(f"{prefix}[{i}]: {val_str}")
    
    format_dict(raw_data)
    return "\n".join(lines)


def format_dadata_api_data(raw_data):
    """
    Форматирует данные из DaData API в текстовый формат "key: val\nkey: val"
    
    Args:
        raw_data: Словарь с данными из DaData API
        
    Returns:
        Строка с отформатированными данными
    """
    if not raw_data:
        return ""
    
    lines = []
    
    def format_value(value):
        """Форматирует простое значение"""
        if value is None:
            return ""
        elif isinstance(value, bool):
            return "Да" if value else "Нет"
        else:
            return str(value)
    
    def format_list(lst, prefix=""):
        """Форматирует список"""
        for i, item in enumerate(lst):
            if isinstance(item, dict):
                lines.append(f"{prefix}[{i}]:")
                format_dict(item, prefix + "  ")
            else:
                val_str = format_value(item)
                if val_str:
                    lines.append(f"{prefix}[{i}]: {val_str}")
    
    def format_dict(d, prefix=""):
        """Форматирует словарь"""
        for key, value in d.items():
            if isinstance(value, dict):
                lines.append(f"{prefix}{key}:")
                format_dict(value, prefix + "  ")
            elif isinstance(value, list):
                if len(value) > 0:
                    lines.append(f"{prefix}{key}:")
                    format_list(value, prefix + "  ")
                else:
                    lines.append(f"{prefix}{key}: []")
            else:
                val_str = format_value(value)
                if val_str:
                    lines.append(f"{prefix}{key}: {val_str}")
    
    format_dict(raw_data)
    return "\n".join(lines)


def send_admin_notification(seller):
    """Отправка уведомления администратору о новой регистрации"""
    try:
        subject = f'Новая регистрация продавца: {seller.company_name}'
        message = f'''
        Новый продавец зарегистрировался на платформе:

        Компания: {seller.company_name}
        Контактное лицо: {seller.contact_person}
        Email: {seller.user.email}
        Телефон: {seller.phone}
        ИНН: {seller.inn}

        Для проверки перейдите в админ-панель.
        '''

        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [admin[1] for admin in settings.ADMINS],  # Отправляем всем админам
            fail_silently=True,
        )
    except Exception as e:
        logger.error(f"Error sending admin notification: {e}")


# API для поиска по ИНН через DaData
@csrf_exempt
def search_company_by_inn(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            inn = data.get('inn', '').strip()

            if not inn or (len(inn) != 10 and len(inn) != 12):
                return JsonResponse({'success': False, 'error': 'Введите корректный ИНН (10 или 12 цифр)'})

            # Используем DaData API для поиска компании
            from sellers.api_services.dadata_service import DaDataService
            
            dadata_service = DaDataService()
            company_data = dadata_service.get_company_by_inn(inn)
            
            if not company_data:
                return JsonResponse({
                    'success': False, 
                    'error': 'Компания с указанным ИНН не найдена. Проверьте правильность ИНН.'
                })
            
            return JsonResponse({
                'success': True,
                'data': company_data
            })

        except json.JSONDecodeError:
            logger.error("Ошибка парсинга JSON в запросе поиска по ИНН")
            return JsonResponse({'success': False, 'error': 'Неверный формат запроса'})
        except Exception as e:
            logger.error(f"Error in INN search: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Ошибка при поиске компании. Попробуйте позже.'})

    return JsonResponse({'success': False, 'error': 'Метод не разрешен'})


# Получение списка категорий для AJAX
def get_categories_json(request):
    categories = Category.objects.filter(is_active=True).values('id', 'name')
    return JsonResponse(list(categories), safe=False)


def get_categories_api(request):
    try:
        categories = Category.objects.all().order_by('name')
        categories_data = list(categories.values('id', 'name', 'slug'))
        return JsonResponse(categories_data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Раздел товаров
@login_required
@ensure_csrf_cookie
def seller_products(request):
    """Страница управления товарами продавца"""
    print("DEBUG: seller_products view called")

    # Временный объект продавца для разработки
    seller = None
    try:
        seller = request.user.seller
    except Seller.DoesNotExist:
        # Создаем временного продавца для отладки
        seller, created = Seller.objects.get_or_create(
            user=request.user,
            defaults={
                'company_name': 'Временная компания',
                'contact_person': request.user.get_full_name() or request.user.username,
                'phone': '+79999999999',
                'inn': '1234567890',
                'status': 'active'
            }
        )
        if created:
            print(f"Создан временный продавец для пользователя {request.user.username}")

    # Базовый queryset товаров продавца
    products = Product.objects.filter(seller=seller).select_related('category').order_by('-created_at')

    # Фильтрация
    search_query = request.GET.get('search', '')
    category_filter = request.GET.get('category', '')
    status_filter = request.GET.get('status', '')
    product_type_filter = request.GET.get('product_type', '')

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(part_number__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    if category_filter:
        products = products.filter(category_id=category_filter)

    if product_type_filter:
        products = products.filter(product_type=product_type_filter)

    if status_filter:
        if status_filter == 'active':
            products = products.filter(is_active=True, stock__gt=0)
        elif status_filter == 'inactive':
            products = products.filter(is_active=False)
        elif status_filter == 'low_stock':
            products = products.filter(is_active=True, stock__lte=10, stock__gt=0)
        elif status_filter == 'out_of_stock':
            products = products.filter(stock=0)

    # Пагинация
    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Статистика для синхронизации с dashboard
    total_products = products.count()
    active_products = products.filter(is_active=True, stock__gt=0).count()
    low_stock_products = products.filter(is_active=True, stock__lte=10, stock__gt=0).count()
    out_of_stock_products = products.filter(stock=0).count()

    # Расчет изменений за неделю для dashboard
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    products_week_ago = Product.objects.filter(
        seller=seller,
        is_active=True,
        stock__gt=0,
        created_at__date__lte=week_ago
    ).count()

    products_change = 0
    if products_week_ago > 0:
        products_change = ((active_products - products_week_ago) / products_week_ago) * 100

    # Получаем категории в иерархическом формате
    hierarchical_categories = get_hierarchical_categories()

    context = {
        'seller': seller,
        'products': page_obj,
        'categories': hierarchical_categories,  # Теперь иерархические категории
        'product_types': Product.PRODUCT_TYPES,
        'total_products': total_products,
        'active_products': active_products,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'products_change': products_change,
    }

    # Если это AJAX запрос, возвращаем только таблицу
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('ajax'):
        return render(request, 'sellers/partials/product_table.html', context)

    return render(request, 'sellers/seller_product.html', context)


# иерархический список
def get_hierarchical_categories():
    """
    Возвращает категории в иерархическом формате для выпадающего списка
    """

    def get_category_tree(parent=None, level=0):
        categories = []
        children = Category.objects.filter(parent=parent, is_active=True).order_by('order', 'name')

        for category in children:
            # Добавляем отступы для вложенности
            prefix = "--- " * level
            categories.append({
                'id': category.id,
                'name': f"{prefix}{category.name}",
                'object': category
            })
            # Рекурсивно добавляем дочерние категории
            categories.extend(get_category_tree(category, level + 1))

        return categories

    return get_category_tree()


@login_required
def product_add(request):
    """Добавление нового товара"""
    if request.method == 'GET':
        try:
            seller = request.user.seller
        except Seller.DoesNotExist:
            return redirect('sellers:become_seller')
        
        # Получаем родительские категории (Легковые, мото, грузовые, спец, шины и диски)
        parent_categories = Category.objects.filter(
            parent__isnull=True,
            is_active=True,
            show_in__in=['cars', 'trucks', 'moto', 'special', 'tires']
        ).order_by('order', 'name')
        
        context = {
            'seller': seller,
            'parent_categories': parent_categories,
        }
        return render(request, 'sellers/add_product.html', context)
    
    elif request.method == 'POST':
        try:
            seller = request.user.seller
        except Seller.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'У вас нет прав продавца'})

        if seller.status != 'approved':
            return JsonResponse({
                'success': False,
                'error': 'Товары могут добавлять только одобренные продавцы. Дождитесь проверки вашей заявки.'
            })

        try:
            # Валидация обязательных полей
            name = request.POST.get('name', '').strip()
            if not name:
                return JsonResponse({'success': False, 'error': 'Название товара обязательно'})
            
            price = request.POST.get('price')
            if not price:
                return JsonResponse({'success': False, 'error': 'Цена обязательна'})
            
            try:
                price = float(price)
                if price <= 0:
                    return JsonResponse({'success': False, 'error': 'Цена должна быть больше 0'})
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'error': 'Некорректная цена'})
            
            # Получаем категорию из shop
            category_id = request.POST.get('category')
            if not category_id:
                return JsonResponse({'success': False, 'error': 'Категория обязательна'})
            
            try:
                category = Category.objects.get(id=category_id, is_active=True)
            except Category.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Категория не найдена'})
            
            # Валидация количества
            try:
                quantity = int(request.POST.get('quantity', 1))
                if quantity < 1:
                    return JsonResponse({'success': False, 'error': 'Количество должно быть не меньше 1'})
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'error': 'Некорректное количество'})

            vin_number, vin_error = _get_validated_product_vin(request.POST)
            if vin_error:
                return JsonResponse({'success': False, 'error': vin_error})

            # Статус модерации: проверенные продавцы (10+ доставленных заказов) публикуют сразу
            can_auto_publish = seller.can_auto_publish_products()
            moderation_status = 'approved' if can_auto_publish else 'pending'
            is_active = can_auto_publish

            # Получаем данные формы
            product = Product(
                seller=seller,
                name=name,
                part_number=request.POST.get('part_number', '').strip() or None,
                category=category,
                product_type=request.POST.get('product_type', 'spare_part'),
                price=price,
                stock=quantity,
                description=request.POST.get('description', '').strip(),
                vin=vin_number,
                old_price=request.POST.get('old_price') or None,
                weight=request.POST.get('weight') or None,
                length=request.POST.get('length') or None,
                width=request.POST.get('width') or None,
                height=request.POST.get('height') or None,
                under_order=request.POST.get('under_order') == 'on',
                country_of_origin=request.POST.get('country_of_origin', '').strip(),
                is_original=request.POST.get('is_original') == 'on',
                moderation_status=moderation_status,
                is_active=is_active
            )
            
            # Сохраняем штрихкод в дополнительных данных, если поле есть
            # Если в модели нет поля barcode, можно сохранить в part_number или создать отдельное поле
            barcode = request.POST.get('barcode', '')
            if barcode and not product.part_number:
                # Если нет артикула, используем штрихкод как артикул
                product.part_number = barcode
            
            # Обработка бренда товара
            brand_name = request.POST.get('brand', '').strip()
            if brand_name:
                from shop.models import Manufacturer
                manufacturer, created = Manufacturer.objects.get_or_create(
                    name=brand_name,
                    defaults={'is_active': True}
                )
                product.manufacturer = manufacturer
            
            # Обработка типа транспорта и связей
            vehicle_type = request.POST.get('vehicle_type')
            if vehicle_type:
                product.vehicle_type = vehicle_type
                
                # Устанавливаем связи с транспортом
                if vehicle_type == 'car':
                    car_brand_id = request.POST.get('vehicle_brand')
                    if car_brand_id:
                        from shop.models import CarBrand
                        try:
                            product.car_brand = CarBrand.objects.get(id=car_brand_id)
                        except CarBrand.DoesNotExist:
                            pass
                    
                    car_model_id = request.POST.get('vehicle_model')
                    if car_model_id:
                        from shop.models import CarModel
                        try:
                            product.save()
                            product.car_models.add(CarModel.objects.get(id=car_model_id))
                        except CarModel.DoesNotExist:
                            pass
                    
                    car_generation_id = request.POST.get('vehicle_generation')
                    if car_generation_id:
                        from shop.models import CarGeneration
                        try:
                            product.save()  # Сохраняем перед добавлением ManyToMany
                            product.car_generations.add(CarGeneration.objects.get(id=car_generation_id))
                        except CarGeneration.DoesNotExist:
                            pass
                    
                    car_modification_id = request.POST.get('vehicle_modification')
                    if car_modification_id:
                        from shop.models import CarModification
                        try:
                            product.save()  # Сохраняем перед добавлением ManyToMany
                            product.car_modifications.add(CarModification.objects.get(id=car_modification_id))
                        except CarModification.DoesNotExist:
                            pass
                            
                elif vehicle_type == 'truck':
                    truck_brand_id = request.POST.get('vehicle_brand')
                    if truck_brand_id:
                        from shop.models import TruckBrand
                        try:
                            product.truck_brand = TruckBrand.objects.get(id=truck_brand_id)
                        except TruckBrand.DoesNotExist:
                            pass
                    
                    truck_model_id = request.POST.get('vehicle_model')
                    if truck_model_id:
                        from shop.models import TruckModel
                        try:
                            product.save()
                            product.truck_models.add(TruckModel.objects.get(id=truck_model_id))
                        except TruckModel.DoesNotExist:
                            pass
                    
                    truck_generation_id = request.POST.get('vehicle_generation')
                    if truck_generation_id:
                        from shop.models import TruckGeneration
                        try:
                            product.save()  # Сохраняем перед добавлением ManyToMany
                            product.truck_generations.add(TruckGeneration.objects.get(id=truck_generation_id))
                        except TruckGeneration.DoesNotExist:
                            pass
                    
                    truck_modification_id = request.POST.get('vehicle_modification')
                    if truck_modification_id:
                        from shop.models import TruckModification
                        try:
                            product.save()  # Сохраняем перед добавлением ManyToMany
                            product.truck_modifications.add(TruckModification.objects.get(id=truck_modification_id))
                        except TruckModification.DoesNotExist:
                            pass
                            
                elif vehicle_type == 'moto':
                    moto_brand_id = request.POST.get('vehicle_brand')
                    if moto_brand_id:
                        from shop.models import MotoBrand
                        try:
                            product.moto_brand = MotoBrand.objects.get(id=moto_brand_id)
                        except MotoBrand.DoesNotExist:
                            pass
                    
                    moto_model_id = request.POST.get('vehicle_model')
                    if moto_model_id:
                        from shop.models import MotoModel
                        try:
                            product.save()
                            product.moto_models.add(MotoModel.objects.get(id=moto_model_id))
                        except MotoModel.DoesNotExist:
                            pass
                    
                    moto_generation_id = request.POST.get('vehicle_generation')
                    if moto_generation_id:
                        from shop.models import MotoGeneration
                        try:
                            product.save()  # Сохраняем перед добавлением ManyToMany
                            product.moto_generations.add(MotoGeneration.objects.get(id=moto_generation_id))
                        except MotoGeneration.DoesNotExist:
                            pass
                    
                    moto_modification_id = request.POST.get('vehicle_modification')
                    if moto_modification_id:
                        from shop.models import MotoModification
                        try:
                            product.save()  # Сохраняем перед добавлением ManyToMany
                            product.moto_modifications.add(MotoModification.objects.get(id=moto_modification_id))
                        except MotoModification.DoesNotExist:
                            pass
                            
                elif vehicle_type == 'special':
                    special_brand_id = request.POST.get('vehicle_brand')
                    if special_brand_id:
                        from shop.models import SpecialBrand
                        try:
                            product.special_brand = SpecialBrand.objects.get(id=special_brand_id)
                        except SpecialBrand.DoesNotExist:
                            pass
                    
                    special_model_id = request.POST.get('vehicle_model')
                    if special_model_id:
                        from shop.models import SpecialModel
                        try:
                            product.save()
                            product.special_models.add(SpecialModel.objects.get(id=special_model_id))
                        except SpecialModel.DoesNotExist:
                            pass
                    
                    special_generation_id = request.POST.get('vehicle_generation')
                    if special_generation_id:
                        from shop.models import SpecialGeneration
                        try:
                            product.save()  # Сохраняем перед добавлением ManyToMany
                            product.special_generations.add(SpecialGeneration.objects.get(id=special_generation_id))
                        except SpecialGeneration.DoesNotExist:
                            pass
                    
                    special_modification_id = request.POST.get('vehicle_modification')
                    if special_modification_id:
                        from shop.models import SpecialModification
                        try:
                            product.save()  # Сохраняем перед добавлением ManyToMany
                            product.special_modifications.add(SpecialModification.objects.get(id=special_modification_id))
                        except SpecialModification.DoesNotExist:
                            pass
            
            product.save()
            
            # Обработка изображений
            image_index = 0
            while f'image_{image_index}' in request.FILES:
                image_file = request.FILES[f'image_{image_index}']
                if image_index == 0:
                    product.image = image_file
                    product.save()
                # Можно добавить обработку дополнительных изображений если есть модель ProductImage
                image_index += 1

            return JsonResponse({
                'success': True,
                'product_id': product.id,
                'message': 'Товар успешно добавлен'
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})


@login_required
def product_edit(request, product_id):
    """Страница редактирования товара (та же форма, что и добавление)."""
    try:
        seller = request.user.seller
    except Seller.DoesNotExist:
        return redirect('sellers:become_seller')

    product = get_object_or_404(Product, id=product_id, seller=seller)

    if request.method == 'GET':
        parent_categories = Category.objects.filter(
            parent__isnull=True,
            is_active=True,
            show_in__in=['cars', 'trucks', 'moto', 'special', 'tires']
        ).order_by('order', 'name')
        category_path = list(product.category.get_full_path()) if product.category else []
        # ID транспорта для предзаполнения (первая модель/поколение/модификация)
        edit_vehicle_brand_id = None
        edit_vehicle_model_id = None
        edit_vehicle_generation_id = None
        edit_vehicle_modification_id = None
        vt = product.vehicle_type or ''
        if vt == 'car':
            if product.car_brand_id:
                edit_vehicle_brand_id = product.car_brand_id
            first_model = product.car_models.first()
            if first_model:
                edit_vehicle_model_id = first_model.id
            first_gen = product.car_generations.first()
            if first_gen:
                edit_vehicle_generation_id = first_gen.id
            first_mod = product.car_modifications.first()
            if first_mod:
                edit_vehicle_modification_id = first_mod.id
        elif vt == 'truck':
            if product.truck_brand_id:
                edit_vehicle_brand_id = product.truck_brand_id
            first_model = product.truck_models.first()
            if first_model:
                edit_vehicle_model_id = first_model.id
            first_gen = product.truck_generations.first()
            if first_gen:
                edit_vehicle_generation_id = first_gen.id
            first_mod = product.truck_modifications.first()
            if first_mod:
                edit_vehicle_modification_id = first_mod.id
        elif vt == 'moto':
            if product.moto_brand_id:
                edit_vehicle_brand_id = product.moto_brand_id
            first_model = product.moto_models.first()
            if first_model:
                edit_vehicle_model_id = first_model.id
            first_gen = product.moto_generations.first()
            if first_gen:
                edit_vehicle_generation_id = first_gen.id
            first_mod = product.moto_modifications.first()
            if first_mod:
                edit_vehicle_modification_id = first_mod.id
        elif vt == 'special':
            if product.special_brand_id:
                edit_vehicle_brand_id = product.special_brand_id
            first_model = product.special_models.first()
            if first_model:
                edit_vehicle_model_id = first_model.id
            first_gen = product.special_generations.first()
            if first_gen:
                edit_vehicle_generation_id = first_gen.id
            first_mod = product.special_modifications.first()
            if first_mod:
                edit_vehicle_modification_id = first_mod.id
        context = {
            'seller': seller,
            'product': product,
            'parent_categories': parent_categories,
            'edit_mode': True,
            'category_path': category_path,
            'edit_vehicle_brand_id': edit_vehicle_brand_id,
            'edit_vehicle_model_id': edit_vehicle_model_id,
            'edit_vehicle_generation_id': edit_vehicle_generation_id,
            'edit_vehicle_modification_id': edit_vehicle_modification_id,
        }
        return render(request, 'sellers/add_product.html', context)

    # POST — обновление товара (логика как в product_add)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})

    try:
        name = request.POST.get('name', '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'Название товара обязательно'})
        price = request.POST.get('price')
        if not price:
            return JsonResponse({'success': False, 'error': 'Цена обязательна'})
        try:
            price = float(price)
            if price <= 0:
                return JsonResponse({'success': False, 'error': 'Цена должна быть больше 0'})
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Некорректная цена'})
        category_id = request.POST.get('category')
        if not category_id:
            return JsonResponse({'success': False, 'error': 'Категория обязательна'})
        try:
            category = Category.objects.get(id=category_id, is_active=True)
        except Category.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Категория не найдена'})
        try:
            quantity = int(request.POST.get('quantity', 1))
            if quantity < 1:
                return JsonResponse({'success': False, 'error': 'Количество должно быть не меньше 1'})
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Некорректное количество'})

        vin_number, vin_error = _get_validated_product_vin(request.POST)
        if vin_error:
            return JsonResponse({'success': False, 'error': vin_error})

        product.name = name
        product.part_number = request.POST.get('part_number', '').strip() or None
        product.category = category
        product.product_type = request.POST.get('product_type', 'spare_part')
        product.price = price
        product.stock = quantity
        product.description = request.POST.get('description', '').strip()
        product.vin = vin_number
        # Приведение числовых полей к Decimal или None
        def _to_decimal(val):
            if val is None or (isinstance(val, str) and not val.strip()):
                return None
            try:
                return Decimal(str(val).strip())
            except (ValueError, TypeError):
                return None
        product.old_price = _to_decimal(request.POST.get('old_price'))
        product.weight = _to_decimal(request.POST.get('weight'))
        product.length = _to_decimal(request.POST.get('length'))
        product.width = _to_decimal(request.POST.get('width'))
        product.height = _to_decimal(request.POST.get('height'))
        product.under_order = request.POST.get('under_order') == 'on'
        product.country_of_origin = request.POST.get('country_of_origin', '').strip()
        product.is_original = request.POST.get('is_original') == 'on'

        barcode = request.POST.get('barcode', '')
        if barcode and not product.part_number:
            product.part_number = barcode

        brand_name = request.POST.get('brand', '').strip()
        if brand_name:
            from shop.models import Manufacturer
            manufacturer, _ = Manufacturer.objects.get_or_create(
                name=brand_name, defaults={'is_active': True}
            )
            product.manufacturer = manufacturer
        else:
            product.manufacturer = None

        vehicle_type = request.POST.get('vehicle_type') or ''
        product.vehicle_type = vehicle_type if vehicle_type else 'universal'
        product.car_brand = None
        product.truck_brand = None
        product.moto_brand = None
        product.special_brand = None
        product.car_models.clear()
        product.car_generations.clear()
        product.car_modifications.clear()
        product.truck_models.clear()
        product.truck_generations.clear()
        product.truck_modifications.clear()
        product.moto_models.clear()
        product.moto_generations.clear()
        product.moto_modifications.clear()
        product.special_models.clear()
        product.special_generations.clear()
        product.special_modifications.clear()

        if vehicle_type == 'car':
            car_brand_id = request.POST.get('vehicle_brand')
            if car_brand_id:
                from shop.models import CarBrand
                try:
                    product.car_brand = CarBrand.objects.get(id=car_brand_id)
                except CarBrand.DoesNotExist:
                    pass
            car_model_id = request.POST.get('vehicle_model')
            if car_model_id:
                from shop.models import CarModel
                try:
                    product.car_models.add(CarModel.objects.get(id=car_model_id))
                except CarModel.DoesNotExist:
                    pass
            car_generation_id = request.POST.get('vehicle_generation')
            if car_generation_id:
                from shop.models import CarGeneration
                try:
                    product.car_generations.add(CarGeneration.objects.get(id=car_generation_id))
                except CarGeneration.DoesNotExist:
                    pass
            car_modification_id = request.POST.get('vehicle_modification')
            if car_modification_id:
                from shop.models import CarModification
                try:
                    product.car_modifications.add(CarModification.objects.get(id=car_modification_id))
                except CarModification.DoesNotExist:
                    pass
        elif vehicle_type == 'truck':
            truck_brand_id = request.POST.get('vehicle_brand')
            if truck_brand_id:
                from shop.models import TruckBrand
                try:
                    product.truck_brand = TruckBrand.objects.get(id=truck_brand_id)
                except TruckBrand.DoesNotExist:
                    pass
            truck_model_id = request.POST.get('vehicle_model')
            if truck_model_id:
                from shop.models import TruckModel
                try:
                    product.truck_models.add(TruckModel.objects.get(id=truck_model_id))
                except TruckModel.DoesNotExist:
                    pass
            truck_generation_id = request.POST.get('vehicle_generation')
            if truck_generation_id:
                from shop.models import TruckGeneration
                try:
                    product.truck_generations.add(TruckGeneration.objects.get(id=truck_generation_id))
                except TruckGeneration.DoesNotExist:
                    pass
            truck_modification_id = request.POST.get('vehicle_modification')
            if truck_modification_id:
                from shop.models import TruckModification
                try:
                    product.truck_modifications.add(TruckModification.objects.get(id=truck_modification_id))
                except TruckModification.DoesNotExist:
                    pass
        elif vehicle_type == 'moto':
            moto_brand_id = request.POST.get('vehicle_brand')
            if moto_brand_id:
                from shop.models import MotoBrand
                try:
                    product.moto_brand = MotoBrand.objects.get(id=moto_brand_id)
                except MotoBrand.DoesNotExist:
                    pass
            moto_model_id = request.POST.get('vehicle_model')
            if moto_model_id:
                from shop.models import MotoModel
                try:
                    product.moto_models.add(MotoModel.objects.get(id=moto_model_id))
                except MotoModel.DoesNotExist:
                    pass
            moto_generation_id = request.POST.get('vehicle_generation')
            if moto_generation_id:
                from shop.models import MotoGeneration
                try:
                    product.moto_generations.add(MotoGeneration.objects.get(id=moto_generation_id))
                except MotoGeneration.DoesNotExist:
                    pass
            moto_modification_id = request.POST.get('vehicle_modification')
            if moto_modification_id:
                from shop.models import MotoModification
                try:
                    product.moto_modifications.add(MotoModification.objects.get(id=moto_modification_id))
                except MotoModification.DoesNotExist:
                    pass
        elif vehicle_type == 'special':
            special_brand_id = request.POST.get('vehicle_brand')
            if special_brand_id:
                from shop.models import SpecialBrand
                try:
                    product.special_brand = SpecialBrand.objects.get(id=special_brand_id)
                except SpecialBrand.DoesNotExist:
                    pass
            special_model_id = request.POST.get('vehicle_model')
            if special_model_id:
                from shop.models import SpecialModel
                try:
                    product.special_models.add(SpecialModel.objects.get(id=special_model_id))
                except SpecialModel.DoesNotExist:
                    pass
            special_generation_id = request.POST.get('vehicle_generation')
            if special_generation_id:
                from shop.models import SpecialGeneration
                try:
                    product.special_generations.add(SpecialGeneration.objects.get(id=special_generation_id))
                except SpecialGeneration.DoesNotExist:
                    pass
            special_modification_id = request.POST.get('vehicle_modification')
            if special_modification_id:
                from shop.models import SpecialModification
                try:
                    product.special_modifications.add(SpecialModification.objects.get(id=special_modification_id))
                except SpecialModification.DoesNotExist:
                    pass

        # После правок по замечаниям модератора — снова на проверку
        if getattr(product, 'moderation_status', None) == 'revision':
            product.moderation_status = 'pending'
            product.moderation_notes = ''  # очищаем после правок

        product.save()

        image_index = 0
        while f'image_{image_index}' in request.FILES:
            image_file = request.FILES[f'image_{image_index}']
            if image_index == 0:
                product.image = image_file
                product.save()
            image_index += 1

        return JsonResponse({
            'success': True,
            'message': 'Товар успешно обновлён',
            'redirect': reverse('sellers:products'),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def product_duplicate(request, product_id):
    """Дублирование товара с полной формой"""
    if request.method == 'POST':
        try:
            seller = request.user.seller
            original_product = get_object_or_404(Product, id=product_id, seller=seller)

            # Получаем данные из формы
            copy_images = request.POST.get('copy_images') == 'true'

            # Создаем копию товара
            duplicated_product = Product(
                seller=seller,
                name=request.POST.get('name'),
                part_number=request.POST.get('part_number', ''),
                category_id=request.POST.get('category'),
                product_type=request.POST.get('product_type', 'spare_part'),
                vin_number=request.POST.get('vin_number', ''),
                barcode=request.POST.get('barcode', ''),
                price=request.POST.get('price'),
                old_price=request.POST.get('old_price') or None,
                quantity=request.POST.get('quantity', 1),
                stock=request.POST.get('stock', 0),
                description=request.POST.get('description', ''),
                specifications=request.POST.get('specifications', ''),
                is_active=request.POST.get('is_active') == 'on',
                is_available=request.POST.get('is_available') == 'on',
                has_guarantee=request.POST.get('has_guarantee') == 'on',
                is_original=request.POST.get('is_original') == 'on'
            )
            duplicated_product.save()

            # Копируем изображение если есть и включена опция
            if copy_images and original_product.image:
                duplicated_product.image.save(
                    original_product.image.name,
                    original_product.image.file,
                    save=True
                )

            # Обрабатываем загруженные изображения
            images = request.FILES.getlist('images')
            for image in images:
                # Здесь можно сохранить дополнительные изображения
                # В зависимости от вашей модели изображений
                pass

            # Получаем обновленную статистику
            products = Product.objects.filter(seller=seller)
            stats = {
                'total_products': products.count(),
                'active_products': products.filter(is_active=True, stock__gt=0).count(),
                'low_stock_products': products.filter(is_active=True, stock__lte=10, stock__gt=0).count(),
                'out_of_stock_products': products.filter(stock=0).count(),
            }

            return JsonResponse({
                'success': True,
                'message': f'Товар продублирован',
                'product_id': duplicated_product.id,
                'stats': stats
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})


def get_recent_products_count(seller, days=7):
    """Получить количество добавленных товаров за указанное количество дней"""
    from django.utils import timezone
    from datetime import timedelta

    date_threshold = timezone.now() - timedelta(days=days)

    recent_products = Product.objects.filter(
        seller=seller,
        created_at__gte=date_threshold
    ).count()

    return recent_products


@login_required
def product_delete(request, product_id):
    """Удаление товара"""
    if request.method == 'POST':
        try:
            seller = request.user.seller
            product = get_object_or_404(Product, id=product_id, seller=seller)
            product_name = product.name
            product.delete()

            # Получаем обновленную статистику после удаления
            products = Product.objects.filter(seller=seller)
            stats = {
                'total_products': products.count(),
                'active_products': products.filter(is_active=True, stock__gt=0).count(),
                'low_stock_products': products.filter(is_active=True, stock__lte=10, stock__gt=0).count(),
                'out_of_stock_products': products.filter(stock=0).count(),
            }

            return JsonResponse({
                'success': True,
                'success': True,
                'message': f'Товар "{product_name}" удален',
                'stats': stats
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})


@login_required
def product_bulk_action(request):
    """Массовые действия с товарами"""
    if request.method == 'POST':
        try:
            seller = request.user.seller
            data = json.loads(request.body)
            action = data.get('action')
            product_ids = data.get('product_ids', [])

            products = Product.objects.filter(id__in=product_ids, seller=seller)

            if action == 'activate':
                products.update(is_active=True)
            elif action == 'deactivate':
                products.update(is_active=False)
            elif action == 'delete':
                products.delete()

            return JsonResponse({'success': True, 'message': f'Действие выполнено для {len(product_ids)} товаров'})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})


@login_required
def seller_orders(request):
    """Страница заказов (временная заглушка)"""
    try:
        seller = request.user.seller
    except Seller.DoesNotExist:
        messages.error(request, 'У вас нет прав продавца')
        return redirect('sellers:dashboard')

    context = {'seller': seller}
    return render(request, 'sellers/seller_orders.html', context)


@login_required
def seller_analytics(request):
    """Страница аналитики (временная заглушка)"""
    try:
        seller = request.user.seller
    except Seller.DoesNotExist:
        messages.error(request, 'У вас нет прав продавца')
        return redirect('sellers:dashboard')

    context = {'seller': seller}
    return render(request, 'sellers/seller_analytics.html', context)


@login_required
def seller_delivery(request):
    """Страница доставки (временная заглушка)"""
    try:
        seller = request.user.seller
    except Seller.DoesNotExist:
        messages.error(request, 'У вас нет прав продавца')
        return redirect('sellers:dashboard')

    context = {'seller': seller}
    return render(request, 'sellers/seller_delivery.html', context)


@login_required
def seller_finances(request):
    """Страница финансов (временная заглушка)"""
    try:
        seller = request.user.seller
    except Seller.DoesNotExist:
        messages.error(request, 'У вас нет прав продавца')
        return redirect('sellers:dashboard')

    context = {'seller': seller}
    return render(request, 'sellers/finances.html', context)


@login_required
def seller_reviews(request):
    """Страница отзывов с полной статистикой"""
    try:
        seller = request.user.seller
    except Seller.DoesNotExist:
        messages.error(request, 'У вас нет прав продавца')
        return redirect('sellers:dashboard')

    # Получаем или создаем сводную статистику
    summary, created = SellerReviewSummary.objects.get_or_create(seller=seller)
    if created:
        summary.update_statistics()

    # Последние отзывы
    recent_reviews = ProductReview.objects.filter(
        product__seller=seller
    ).select_related('product', 'user').prefetch_related('media_files').order_by('-created_at')[:10]

    # Отзывы требующие внимания
    attention_reviews = ProductReview.objects.filter(
        product__seller=seller,
        status='approved',
        seller_response=''
    ).select_related('product', 'user')[:5]

    # Статистика по статусам
    status_stats = ProductReview.objects.filter(
        product__seller=seller
    ).values('status').annotate(count=Count('id'))

    total_status = sum(stat['count'] for stat in status_stats)
    for stat in status_stats:
        stat['percentage'] = (stat['count'] / total_status * 100) if total_status > 0 else 0

    # Создаем словарь с данными summary для шаблона
    review_summary_dict = {
        'total_reviews': summary.total_reviews,
        'average_rating': float(summary.average_rating) if summary.average_rating else 0,
        'response_rate': float(summary.response_rate) if summary.response_rate else 0,
        'verified_reviews': summary.verified_reviews,
        'rating_1': summary.rating_1,
        'rating_2': summary.rating_2,
        'rating_3': summary.rating_3,
        'rating_4': summary.rating_4,
        'rating_5': summary.rating_5,
        'get_rating_percentage': summary.get_rating_percentage() if hasattr(summary, 'get_rating_percentage') else {
            1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    }

    context = {
        'seller': seller,
        'review_summary': review_summary_dict,
        'recent_reviews': recent_reviews,
        'attention_reviews': attention_reviews,
        'status_stats': status_stats,
    }

    return render(request, 'sellers/seller_reviews.html', context)


@login_required
def seller_settings(request):
    """Страница настроек (временная заглушка)"""
    try:
        seller = request.user.seller
    except Seller.DoesNotExist:
        messages.error(request, 'У вас нет прав продавца')
        return redirect('sellers:dashboard')

    context = {'seller': seller}
    return render(request, 'sellers/seller_settings.html', context)


@login_required
def seller_help(request):
    """Страница помощи (временная заглушка)"""
    try:
        seller = request.user.seller
    except Seller.DoesNotExist:
        messages.error(request, 'У вас нет прав продавца')
        return redirect('sellers:dashboard')

    context = {'seller': seller}
    return render(request, 'sellers/seller_help.html', context)


@login_required
def product_import(request):
    """Импорт товаров из файла"""
    if request.method == 'POST':
        try:
            seller = request.user.seller
            uploaded_file = request.FILES['file']
            update_existing = request.POST.get('update_existing') == 'true'
            deactivate_missing = request.POST.get('deactivate_missing') == 'true'

            # Здесь логика обработки файла
            # processed = process_import_file(uploaded_file, seller, update_existing, deactivate_missing)

            return JsonResponse({
                'success': True,
                'processed': 0,  # Замените на реальное число
                'message': 'Импорт завершен успешно'
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})


@login_required
def product_export(request):
    """Экспорт товаров в файл"""
    format_type = request.GET.get('format', 'csv')
    scope = request.GET.get('scope', 'all')
    fields = request.GET.get('fields', '').split(',')

    try:
        seller = request.user.seller
        products = Product.objects.filter(seller=seller)

        # Применяем фильтры в зависимости от scope
        if scope == 'filtered':
            # Применяем текущие фильтры из запроса
            pass
        elif scope == 'selected':
            # Экспорт только выбранных товаров
            pass
        elif scope == 'active':
            products = products.filter(is_active=True)
        elif scope == 'inactive':
            products = products.filter(is_active=False)

        # Генерируем файл в нужном формате
        response = generate_export_file(products, format_type, fields)
        return response

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def generate_export_file(products, format_type, fields):
    """Генерация файла экспорта"""
    # Реализация генерации файла в нужном формате
    # Возвращает HttpResponse с файлом
    pass


############################### ЗАКАЗЫ ###############################


@login_required
def seller_orders(request):
    """Страница управления заказами продавца"""
    seller = request.user.seller

    # Получаем заказы, которые содержат товары продавца
    orders = Order.objects.filter(
        items__product__seller=seller
    ).distinct().prefetch_related(
        'items',
        'items__product',
        'payments'
    ).order_by('-created_at')

    # Фильтрация
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    paid_filter = request.GET.get('paid', '')

    if search_query:
        orders = orders.filter(
            Q(id__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )

    if status_filter:
        orders = orders.filter(status=status_filter)

    if paid_filter:
        if paid_filter == 'paid':
            orders = orders.filter(paid=True)
        elif paid_filter == 'unpaid':
            orders = orders.filter(paid=False)

    if date_from:
        orders = orders.filter(created_at__gte=date_from)

    if date_to:
        orders = orders.filter(created_at__lte=date_to)

    # Аннотируем заказы информацией о товарах продавца
    orders = orders.annotate(
        seller_items_count=Count('items', filter=Q(items__product__seller=seller)),
        seller_total=Sum(
            F('items__price') * F('items__quantity'),
            filter=Q(items__product__seller=seller)
        )
    )

    # Пагинация
    paginator = Paginator(orders, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Статистика
    total_orders = orders.count()
    processing_orders = orders.filter(status='processing').count()
    shipped_orders = orders.filter(status='shipped').count()
    delivered_orders = orders.filter(status='delivered').count()
    canceled_orders = orders.filter(status='canceled').count()

    # Новые заказы за сегодня (для синхронизации с dashboard)
    today = timezone.now().date()
    today_orders = Order.objects.filter(
        items__product__seller=seller,
        created_at__date=today
    ).distinct().count()

    # Сумма продаж
    total_sales = orders.aggregate(
        total=Sum('seller_total')
    )['total'] or 0

    context = {
        'seller': seller,
        'orders': page_obj,
        'statuses': Order.STATUS_CHOICES,
        'total_orders': total_orders,
        'processing_orders': processing_orders,
        'shipped_orders': shipped_orders,
        'delivered_orders': delivered_orders,
        'canceled_orders': canceled_orders,
        'today_orders': today_orders,  # Добавляем для синхронизации
        'total_sales': total_sales,
    }

    return render(request, 'sellers/seller_orders.html', context)


@login_required
def order_detail(request, order_id):
    """Детальная страница заказа"""
    seller = request.user.seller
    order = get_object_or_404(Order, id=order_id)

    # Проверяем, что в заказе есть товары продавца
    seller_items = order.items.filter(product__seller=seller)
    if not seller_items.exists():
        return JsonResponse({'success': False, 'error': 'Заказ не найден'})

    # Товары продавца в заказе
    seller_order_total = sum(item.get_cost() for item in seller_items)

    context = {
        'seller': seller,
        'order': order,
        'seller_items': seller_items,
        'seller_order_total': seller_order_total,
    }

    return render(request, 'sellers/order_detail.html', context)


@login_required
def update_order_status(request, order_id):
    """Обновление статуса заказа"""
    if request.method == 'POST':
        try:
            seller = request.user.seller
            order = get_object_or_404(Order, id=order_id)

            # Проверяем, что в заказе есть товары продавца
            if not order.items.filter(product__seller=seller).exists():
                return JsonResponse({'success': False, 'error': 'Доступ запрещен'})

            new_status = request.POST.get('status')
            if new_status not in dict(Order.STATUS_CHOICES):
                return JsonResponse({'success': False, 'error': 'Неверный статус'})

            order.status = new_status
            order.save()

            return JsonResponse({
                'success': True,
                'message': f'Статус заказа #{order_id} обновлен',
                'new_status': order.get_status_display(),
                'status_class': new_status
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})


@login_required
def bulk_order_action(request):
    """Массовые действия с заказами"""
    if request.method == 'POST':
        try:
            seller = request.user.seller
            data = json.loads(request.body)
            action = data.get('action')
            order_ids = data.get('order_ids', [])

            # Фильтруем заказы, которые содержат товары продавца
            orders = Order.objects.filter(
                id__in=order_ids,
                items__product__seller=seller
            ).distinct()

            if action == 'mark_processing':
                orders.update(status='processing')
            elif action == 'mark_shipped':
                orders.update(status='shipped')
            elif action == 'mark_delivered':
                orders.update(status='delivered')
            elif action == 'mark_canceled':
                orders.update(status='canceled')

            return JsonResponse({
                'success': True,
                'message': f'Действие выполнено для {orders.count()} заказов'
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})


def get_new_orders_count(seller, hours=24):
    """Получить количество новых заказов за указанное количество часов"""
    from django.utils import timezone
    from datetime import timedelta

    time_threshold = timezone.now() - timedelta(hours=hours)

    new_orders = Order.objects.filter(
        items__product__seller=seller,
        created_at__gte=time_threshold,
        status='processing'  # или другие статусы для "новых" заказов
    ).distinct().count()

    return new_orders


# Экспорт заказов
@login_required
def orders_export(request):
    """Экспорт заказов в различные форматы"""
    seller = request.user.seller

    # Получаем параметры экспорта
    format_type = request.GET.get('format', 'xlsx')
    scope = request.GET.get('scope', 'all')
    fields = request.GET.get('fields', '').split(',')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Базовый queryset
    orders = Order.objects.filter(
        items__product__seller=seller
    ).distinct().prefetch_related(
        'items',
        'items__product'
    ).order_by('-created_at')

    # Применяем фильтры в зависимости от scope
    if scope == 'filtered':
        # Применяем текущие фильтры из запроса
        search_query = request.GET.get('filter_search')
        status_filter = request.GET.get('filter_status')
        paid_filter = request.GET.get('filter_paid')

        if search_query:
            orders = orders.filter(
                Q(id__icontains=search_query) |
                Q(name__icontains=search_query) |
                Q(email__icontains=search_query)
            )

        if status_filter:
            orders = orders.filter(status=status_filter)

        if paid_filter:
            if paid_filter == 'paid':
                orders = orders.filter(paid=True)
            elif paid_filter == 'unpaid':
                orders = orders.filter(paid=False)

    elif scope == 'selected':
        order_ids = request.GET.get('order_ids', '').split(',')
        orders = orders.filter(id__in=order_ids)

    elif scope == 'processing':
        orders = orders.filter(status='processing')
    elif scope == 'shipped':
        orders = orders.filter(status='shipped')
    elif scope == 'delivered':
        orders = orders.filter(status='delivered')
    elif scope == 'canceled':
        orders = orders.filter(status='canceled')
    elif scope == 'paid':
        orders = orders.filter(paid=True)
    elif scope == 'unpaid':
        orders = orders.filter(paid=False)

    # Фильтр по дате
    if date_from:
        orders = orders.filter(created_at__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__lte=date_to)

    # Аннотируем заказы
    orders = orders.annotate(
        seller_items_count=Count('items', filter=Q(items__product__seller=seller)),
        seller_total=Sum(
            F('items__price') * F('items__quantity'),
            filter=Q(items__product__seller=seller)
        )
    )

    # Подготавливаем данные для экспорта
    export_data = []
    for order in orders:
        order_data = {}

        if not fields or 'id' in fields:
            order_data['id'] = order.id
        if not fields or 'created_at' in fields:
            order_data['created_at'] = order.created_at.strftime("%d.%m.%Y %H:%M")
        if not fields or 'customer_name' in fields:
            order_data['customer_name'] = order.name
        if not fields or 'email' in fields:
            order_data['email'] = order.email
        if not fields or 'phone' in fields:
            order_data['phone'] = order.phone
        if not fields or 'address' in fields:
            order_data['address'] = order.address
        if not fields or 'items_count' in fields:
            order_data['items_count'] = order.seller_items_count
        if not fields or 'total_amount' in fields:
            order_data['total_amount'] = float(order.seller_total or 0)
        if not fields or 'status' in fields:
            order_data['status'] = order.get_status_display()
        if not fields or 'paid' in fields:
            order_data['paid'] = 'Да' if order.paid else 'Нет'
        if not fields or 'payment_method' in fields:
            order_data['payment_method'] = order.get_payment_method_display() if order.payment_method else 'Не указан'
        if not fields or 'comment' in fields:
            order_data['comment'] = order.comment or ''
        if not fields or 'products' in fields:
            products = []
            for item in order.items.filter(product__seller=seller):
                products.append(f"{item.product.name} (x{item.quantity})")
            order_data['products'] = "; ".join(products)

        export_data.append(order_data)

    # Генерируем файл в нужном формате
    filename = f"orders_export_{seller.company_name}_{timezone.now().strftime('%Y%m%d_%H%M')}"

    if format_type == 'csv':
        return export_to_csv(export_data, filename)
    elif format_type == 'xlsx':
        return export_to_xlsx(export_data, filename)
    elif format_type == 'pdf':
        return export_to_pdf(export_data, filename, seller)
    else:
        return HttpResponse("Unsupported format", status=400)


def export_to_csv(data, filename):
    """Экспорт в CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

    if not data:
        return response

    writer = csv.writer(response)

    # Заголовки
    writer.writerow(data[0].keys())

    # Данные
    for row in data:
        writer.writerow(row.values())

    return response


def export_to_xlsx(data, filename):
    """Экспорт в Excel"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("Excel export requires openpyxl", status=500)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы"

    if not data:
        wb.save(response)
        return response

    # Заголовки
    headers = list(data[0].keys())
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws.column_dimensions[col_letter].width = 15

    # Данные
    for row_num, row_data in enumerate(data, 2):
        for col_num, key in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=row_data.get(key, ''))

    wb.save(response)
    return response


def export_to_pdf(data, filename, seller):
    """Экспорт в PDF"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        import io
    except ImportError:
        return HttpResponse("PDF export requires reportlab", status=500)

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Заголовок
    p.setFont("Helvetica-Bold", 16)
    p.drawString(20 * mm, height - 20 * mm, f"Отчет по заказам - {seller.company_name}")
    p.setFont("Helvetica", 10)
    p.drawString(20 * mm, height - 25 * mm, f"Сгенерировано: {timezone.now().strftime('%d.%m.%Y %H:%M')}")

    # Данные
    y_position = height - 40 * mm
    for i, order in enumerate(data, 1):
        if y_position < 50 * mm:
            p.showPage()
            y_position = height - 40 * mm

        p.setFont("Helvetica-Bold", 12)
        p.drawString(20 * mm, y_position, f"Заказ #{order.get('id', '')}")
        y_position -= 6 * mm

        p.setFont("Helvetica", 10)
        for key, value in order.items():
            if key != 'id':
                p.drawString(25 * mm, y_position, f"{key}: {value}")
                y_position -= 4 * mm

        y_position -= 4 * mm

    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response


@login_required
def orders_analytics(request):
    """API для аналитики заказов"""
    seller = request.user.seller

    # Получаем параметры периода
    days = int(request.GET.get('days', 30))
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Определяем период
    if date_from and date_to:
        start_date = datetime.strptime(date_from, '%Y-%m-%d')
        end_date = datetime.strptime(date_to, '%Y-%m-%d')
    else:
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)

    # Базовый queryset для периода
    orders = Order.objects.filter(
        items__product__seller=seller,
        created_at__range=[start_date, end_date]
    ).distinct()

    # Предыдущий период для сравнения
    prev_start_date = start_date - (end_date - start_date)
    prev_orders = Order.objects.filter(
        items__product__seller=seller,
        created_at__range=[prev_start_date, start_date]
    ).distinct()

    # Ключевые метрики
    current_metrics = orders.aggregate(
        total_revenue=Sum('seller_total'),
        total_orders=Count('id'),
        avg_order_value=Avg('seller_total')
    )

    prev_metrics = prev_orders.aggregate(
        total_revenue=Sum('seller_total'),
        total_orders=Count('id'),
        avg_order_value=Avg('seller_total')
    )

    # Расчет изменений
    def calculate_change(current, previous):
        if previous and previous > 0:
            return round(((current or 0) - previous) / previous * 100, 1)
        return 0

    metrics = {
        'total_revenue': current_metrics['total_revenue'] or 0,
        'total_orders': current_metrics['total_orders'] or 0,
        'avg_order_value': current_metrics['avg_order_value'] or 0,
        'conversion_rate': 2.5,  # Здесь должна быть реальная логика конверсии
        'revenue_change': calculate_change(
            current_metrics['total_revenue'],
            prev_metrics['total_revenue']
        ),
        'orders_change': calculate_change(
            current_metrics['total_orders'],
            prev_metrics['total_orders']
        ),
        'aov_change': calculate_change(
            current_metrics['avg_order_value'],
            prev_metrics['avg_order_value']
        ),
        'conversion_change': 1.2
    }

    # Данные для графиков
    charts = {
        'sales': get_sales_chart_data(orders, start_date, end_date),
        'status': get_status_chart_data(orders),
        'products': get_products_chart_data(orders),
        'time_distribution': get_time_distribution_data(orders)
    }

    # Дополнительная статистика
    stats = {
        'completed_orders': orders.filter(status='delivered').count(),
        'canceled_orders': orders.filter(status='canceled').count(),
        'refund_rate': 1.2,  # Здесь должна быть логика возвратов
        'repeat_customers': orders.values('email').annotate(
            count=Count('id')
        ).filter(count__gt=1).count()
    }

    return JsonResponse({
        'success': True,
        'metrics': metrics,
        'charts': charts,
        'stats': stats
    })


def get_sales_chart_data(orders, start_date, end_date):
    """Данные для графика продаж по дням"""
    days_diff = (end_date - start_date).days
    dates = [start_date + timedelta(days=i) for i in range(days_diff + 1)]

    daily_data = orders.extra({
        'date': "DATE(created_at)"
    }).values('date').annotate(
        revenue=Sum('seller_total'),
        orders_count=Count('id')
    ).order_by('date')

    # Создаем полный список дат
    revenue_data = [0] * len(dates)
    orders_data = [0] * len(dates)
    labels = []

    for i, date in enumerate(dates):
        labels.append(date.strftime('%d.%m'))
        for data in daily_data:
            if data['date'] == date.date():
                revenue_data[i] = float(data['revenue'] or 0)
                orders_data[i] = data['orders_count']
                break

    return {
        'labels': labels,
        'revenue': revenue_data,
        'orders': orders_data
    }


def get_status_chart_data(orders):
    """Данные для графика статусов заказов"""
    status_data = orders.values('status').annotate(
        count=Count('id')
    ).order_by('status')

    status_labels = {
        'processing': 'В обработке',
        'shipped': 'Отправлены',
        'delivered': 'Доставлены',
        'canceled': 'Отменены'
    }

    labels = []
    data = []

    for item in status_data:
        labels.append(status_labels.get(item['status'], item['status']))
        data.append(item['count'])

    return {
        'labels': labels,
        'data': data
    }


def get_products_chart_data(orders):
    """Данные для топа товаров"""
    from django.db.models import Sum

    product_data = OrderItem.objects.filter(
        order__in=orders,
        product__seller=orders[0].items.first().product.seller if orders.exists() else None
    ).values(
        'product__name'
    ).annotate(
        revenue=Sum(F('price') * F('quantity'))
    ).order_by('-revenue')[:10]

    labels = [item['product__name'][:20] + '...' for item in product_data]
    revenue = [float(item['revenue'] or 0) for item in product_data]

    return {
        'labels': labels,
        'revenue': revenue
    }


def get_time_distribution_data(orders):
    """Распределение заказов по времени суток"""
    hours_data = orders.extra({
        'hour': "EXTRACT(HOUR FROM created_at)"
    }).values('hour').annotate(
        count=Count('id')
    ).order_by('hour')

    # Создаем полный список часов
    hours = list(range(24))
    data = [0] * 24
    labels = [f"{h:02d}:00" for h in hours]

    for item in hours_data:
        hour = int(item['hour'])
        data[hour] = item['count']

    return {
        'labels': labels,
        'data': data
    }


#######АНАЛИТИКА##############


class SellerAdvancedAnalyticsView(TemplateView):
    template_name = 'sellers/advanced_analytics.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        seller = self.request.user

        # Инициализация алгоритмов
        self.sales_predictor = SalesPredictor()
        self.ltv_analyzer = CustomerLTV()
        self.anomaly_detector = AnomalyDetector()
        self.pricing_optimizer = DynamicPricing()
        self.product_segmenter = ProductSegmentation()
        self.marketing_analyzer = MarketingAnalyzer()
        self.alert_system = SmartAlertSystem()
        self.behavior_analyzer = BehaviorAnalyzer()

        # Основные данные
        context.update(self.get_basic_analytics(seller))

        # Продвинутая аналитика
        context.update({
            'sales_predictions': self.get_sales_predictions(seller),
            'customer_segments': self.get_customer_segments(seller),
            'anomalies': self.get_anomalies(seller),
            'pricing_recommendations': self.get_pricing_recommendations(seller),
            'inventory_optimization': self.get_inventory_optimization(seller),
            'marketing_insights': self.get_marketing_insights(seller),
            'smart_alerts': self.get_smart_alerts(seller),
            'behavior_insights': self.get_behavior_insights(seller),
        })

        return context

    def get_sales_predictions(self, seller):
        """Прогнозирование продаж"""
        return self.sales_predictor.predict(
            seller=seller,
            periods=30,
            confidence=0.95
        )

    def get_customer_segments(self, seller):
        """Сегментация клиентов по LTV"""
        return self.ltv_analyzer.analyze_customers(seller)

    def get_anomalies(self, seller):
        """Обнаружение аномалий"""
        return self.anomaly_detector.detect_anomalies(seller)

    def get_pricing_recommendations(self, seller):
        """Рекомендации по ценам"""
        return self.pricing_optimizer.get_recommendations(seller)

    def get_inventory_optimization(self, seller):
        """Оптимизация запасов"""
        return self.product_segmenter.abc_xyz_analysis(seller)

    def get_marketing_insights(self, seller):
        """Анализ эффективности маркетинга"""
        return self.marketing_analyzer.get_roi_analysis(seller)

    def get_smart_alerts(self, seller):
        """Умные уведомления"""
        return self.alert_system.generate_alerts(seller)

    def get_behavior_insights(self, seller):
        """Анализ поведения пользователей"""
        return self.behavior_analyzer.analyze_behavior(seller)


# API endpoints для обновления данных
class AnalyticsDataView(View):
    def get(self, request):
        data_type = request.GET.get('type')
        seller = request.user

        if data_type == 'sales_forecast':
            data = SalesPredictor().get_forecast_data(seller)
        elif data_type == 'customer_segments':
            data = CustomerLTV().get_segment_data(seller)
        elif data_type == 'anomalies':
            data = AnomalyDetector().get_anomaly_data(seller)
        elif data_type == 'pricing':
            data = DynamicPricing().get_pricing_data(seller)

        return JsonResponse(data)


class RunMLAlgorithmView(View):
    def post(self, request):
        algorithm = request.POST.get('algorithm')
        seller = request.user

        if algorithm == 'retrain_forecast':
            result = SalesPredictor().retrain_model(seller)
        elif algorithm == 'optimize_prices':
            result = DynamicPricing().run_optimization(seller)
        elif algorithm == 'segment_customers':
            result = CustomerLTV().update_segments(seller)

        return JsonResponse({'status': 'success', 'result': result})


#### Финансы ####

class FinanceDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'sellers/finances/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        seller = self.request.user.seller

        # Основные финансовые метрики
        finance_profile = seller.sellerfinanceprofile

        # Статистика за разные периоды
        today = datetime.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)

        # Продажи за сегодня
        today_sales = Transaction.objects.filter(
            seller=seller,
            transaction_type='sale',
            transaction_date__date=today,
            status='completed'
        ).aggregate(total=Sum('net_amount'))['total'] or 0

        # Продажи за неделю
        week_sales = Transaction.objects.filter(
            seller=seller,
            transaction_type='sale',
            transaction_date__date__gte=week_ago,
            status='completed'
        ).aggregate(total=Sum('net_amount'))['total'] or 0

        # Продажи за месяц
        month_sales = Transaction.objects.filter(
            seller=seller,
            transaction_type='sale',
            transaction_date__date__gte=month_ago,
            status='completed'
        ).aggregate(total=Sum('net_amount'))['total'] or 0

        # Ожидаемые выплаты
        pending_payouts = Payout.objects.filter(
            seller=seller,
            status__in=['pending', 'processing']
        ).aggregate(total=Sum('amount'))['total'] or 0

        # Последние транзакции
        recent_transactions = Transaction.objects.filter(
            seller=seller
        ).select_related('order')[:10]

        # График доходов за последние 30 дней
        revenue_data = self.get_revenue_chart_data(seller, 30)

        context.update({
            'finance_profile': finance_profile,
            'today_sales': today_sales,
            'week_sales': week_sales,
            'month_sales': month_sales,
            'pending_payouts': pending_payouts,
            'recent_transactions': recent_transactions,
            'revenue_data': revenue_data,
            'seller': seller,
        })
        return context

    def get_revenue_chart_data(self, seller, days=30):
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=days - 1)

        daily_revenue = Transaction.objects.filter(
            seller=seller,
            transaction_type='sale',
            status='completed',
            transaction_date__date__range=[start_date, end_date]
        ).values('transaction_date__date').annotate(
            total_revenue=Sum('net_amount')
        ).order_by('transaction_date__date')

        dates = []
        revenues = []

        # Заполняем все даты в периоде
        current_date = start_date
        while current_date <= end_date:
            dates.append(current_date.strftime('%d.%m'))
            day_revenue = next(
                (item['total_revenue'] for item in daily_revenue
                 if item['transaction_date__date'] == current_date), 0
            )
            revenues.append(float(day_revenue))
            current_date += timedelta(days=1)

        return {
            'labels': dates,
            'datasets': [{
                'label': 'Доход',
                'data': revenues,
                'borderColor': '#667eea',
                'backgroundColor': 'rgba(102, 126, 234, 0.1)',
            }]
        }


class TransactionListView(LoginRequiredMixin, ListView):
    template_name = 'sellers/finances/transactions.html'
    context_object_name = 'transactions'
    paginate_by = 20

    def get_queryset(self):
        seller = self.request.user.seller  # Исправлено
        queryset = Transaction.objects.filter(seller=seller)

        # Фильтрация
        transaction_type = self.request.GET.get('type')
        status = self.request.GET.get('status')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        if transaction_type:
            queryset = queryset.filter(transaction_type=transaction_type)
        if status:
            queryset = queryset.filter(status=status)
        if date_from:
            queryset = queryset.filter(transaction_date__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(transaction_date__date__lte=date_to)

        return queryset.select_related('order')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        seller = self.request.user.seller  # Исправлено

        # Добавляем статистику
        transactions = self.get_queryset()
        context['total_count'] = transactions.count()
        context['total_income'] = transactions.filter(
            transaction_type='sale',
            status='completed'
        ).aggregate(total=Sum('net_amount'))['total'] or 0
        context['total_refunds'] = abs(transactions.filter(
            transaction_type='refund'
        ).aggregate(total=Sum('net_amount'))['total'] or 0)
        context['total_commission'] = transactions.aggregate(
            total=Sum('commission')
        )['total'] or 0
        context['seller'] = seller

        return context


class PayoutRequestView(LoginRequiredMixin, View):
    def post(self, request):
        seller = request.user.seller
        finance_profile = seller.sellerfinanceprofile

        data = json.loads(request.body)
        amount = Decimal(data.get('amount'))
        payout_method = data.get('payout_method')
        account_details = data.get('account_details')

        # Проверка доступного баланса
        if amount > finance_profile.available_balance():
            return JsonResponse({
                'success': False,
                'error': 'Недостаточно средств на балансе'
            })

        # Создание заявки на выплату
        payout = Payout.objects.create(
            seller=seller,
            amount=amount,
            fee=self.calculate_payout_fee(amount, payout_method),
            net_amount=amount - self.calculate_payout_fee(amount, payout_method),
            payout_method=payout_method,
            account_details=account_details,
            payout_reference=f"PO_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{seller.id}"
        )

        # Резервирование средств
        finance_profile.hold_amount += amount
        finance_profile.save()

        return JsonResponse({
            'success': True,
            'message': 'Заявка на выплату создана',
            'payout_id': payout.id
        })

    def calculate_payout_fee(self, amount, method):
        # Логика расчета комиссии в зависимости от способа выплаты
        fees = {
            'bank_card': Decimal('0.02'),  # 2%
            'bank_account': Decimal('0.015'),  # 1.5%
            'yoomoney': Decimal('0.025'),  # 2.5%
            'qiwi': Decimal('0.03'),  # 3%
        }
        return amount * fees.get(method, Decimal('0.02'))


class FinancialReportsView(LoginRequiredMixin, TemplateView):
    template_name = 'sellers/finances/reports.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        seller = self.request.user.seller

        reports = FinancialReport.objects.filter(seller=seller).order_by('-period_end')

        # Статистика для дашборда отчетов
        current_year = datetime.now().year
        yearly_reports = reports.filter(
            report_type='yearly',
            period_start__year=current_year
        )

        context.update({
            'reports': reports,
            'yearly_reports': yearly_reports,
            'current_year': current_year,
            'report_periods': self.get_report_periods(),
            'seller': seller,
        })
        return context

    def get_report_periods(self):
        """Генерируем список периодов для отчетов"""
        today = datetime.now().date()
        periods = []

        # Месячные отчеты за последние 6 месяцев
        for i in range(6):
            month = today.replace(day=1) - timedelta(days=30 * i)
            periods.append({
                'type': 'monthly',
                'start': month.replace(day=1),
                'end': (month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1),
                'name': month.strftime('%B %Y')
            })

        return periods


class GenerateReportView(LoginRequiredMixin, View):
    def post(self, request):
        seller = request.user.seller
        data = json.loads(request.body)

        report_type = data.get('report_type')
        period_start = data.get('period_start')
        period_end = data.get('period_end')
        format_type = data.get('format', 'pdf')

        try:
            # Генерация отчета
            report = self.generate_financial_report(
                seller, report_type, period_start, period_end, format_type
            )

            return JsonResponse({
                'success': True,
                'report_id': report.id,
                'download_url': report.report_file.url if report.report_file else None,
                'message': f'Отчет за {period_start} - {period_end} успешно сгенерирован'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })

    def generate_financial_report(self, seller, report_type, period_start, period_end, format_type):
        """Генерация финансового отчета"""
        from django.utils import timezone
        from decimal import Decimal

        # Расчет финансовых показателей
        transactions = Transaction.objects.filter(
            seller=seller,
            transaction_date__date__range=[period_start, period_end],
            status='completed'
        )

        total_sales = transactions.filter(transaction_type='sale').aggregate(
            total=Sum('net_amount')
        )['total'] or Decimal('0')

        total_refunds = transactions.filter(transaction_type='refund').aggregate(
            total=Sum('net_amount')
        )['total'] or Decimal('0')

        total_commission = transactions.aggregate(
            total=Sum('commission')
        )['total'] or Decimal('0')

        total_payouts = transactions.filter(transaction_type='payout').aggregate(
            total=Sum('net_amount')
        )['total'] or Decimal('0')

        net_profit = total_sales + total_refunds + total_payouts

        # Создание записи отчета
        report = FinancialReport.objects.create(
            seller=seller,
            report_type=report_type,
            period_start=period_start,
            period_end=period_end,
            total_sales=total_sales,
            total_refunds=total_refunds,
            total_commission=total_commission,
            total_payouts=total_payouts,
            net_profit=net_profit,
            tax_amount=net_profit * Decimal('0.06')  # Упрощенный расчет налогов 6%
        )

        # Генерация файла отчета
        report_file = self.generate_report_file(report, format_type)
        if report_file:
            report.report_file = report_file
            report.save()

        return report

    def generate_report_file(self, report, format_type):
        """Генерация файла отчета в выбранном формате"""
        # Здесь будет логика генерации PDF/Excel файлов
        # Пока возвращаем None - файл генерируется по требованию
        return None


class DownloadReportView(LoginRequiredMixin, View):
    def get(self, request, report_id):
        report = get_object_or_404(FinancialReport, id=report_id, seller=request.user.seller)

        if report.report_file:
            return FileResponse(report.report_file.open(), filename=f"report_{report.id}.pdf")
        else:
            # Генерация файла на лету
            return self.generate_on_demand_report(report)

    def generate_on_demand_report(self, report):
        """Генерация отчета по требованию"""
        # Логика генерации PDF
        pass


class ReportDetailsView(LoginRequiredMixin, View):
    def get(self, request, report_id):
        seller = request.user.seller  # ИСПРАВЛЕНО
        report = get_object_or_404(FinancialReport, id=report_id, seller=seller)

        # Детальная статистика для отчета
        transactions = Transaction.objects.filter(
            seller=seller,  # ИСПРАВЛЕНО
            transaction_date__date__range=[report.period_start, report.period_end]
        )

        context = {
            'report': report,
            'transactions_count': transactions.count(),
            'seller': seller,  # ДОБАВЛЕНО
        }

        return render(request, 'sellers/finances/report_details.html', context)


class DeleteReportView(LoginRequiredMixin, View):
    def delete(self, request, report_id):
        report = get_object_or_404(FinancialReport, id=report_id, seller=request.user.seller)
        report.delete()

        return JsonResponse({'success': True, 'message': 'Отчет успешно удален'})


##### Отзывы ###################


class ReviewsDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'sellers/reviews/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        seller = self.request.user.seller

        # Получаем или создаем сводную статистику
        summary, created = SellerReviewSummary.objects.get_or_create(seller=seller)
        if created:
            summary.update_statistics()

        # Последние отзывы
        recent_reviews = ProductReview.objects.filter(
            product__seller=seller
        ).select_related('product', 'user').prefetch_related('media_files').order_by('-created_at')[:10]

        # Отзывы требующие внимания
        attention_reviews = ProductReview.objects.filter(
            product__seller=seller,
            status='approved',
            seller_response=''
        ).select_related('product', 'user')[:5]

        # Статистика по статусам
        status_stats = self.get_status_stats(seller)

        # Создаем дефолтные значения для review_summary если что-то пошло не так
        default_summary = {
            'total_reviews': 0,
            'average_rating': 0,
            'response_rate': 0,
            'verified_reviews': 0,
            'rating_1': 0,
            'rating_2': 0,
            'rating_3': 0,
            'rating_4': 0,
            'rating_5': 0,
        }

        # Обновляем дефолтные значения реальными данными
        if hasattr(summary, 'total_reviews'):
            default_summary = {
                'total_reviews': summary.total_reviews,
                'average_rating': float(summary.average_rating) if summary.average_rating else 0,
                'response_rate': float(summary.response_rate) if summary.response_rate else 0,
                'verified_reviews': summary.verified_reviews,
                'rating_1': summary.rating_1,
                'rating_2': summary.rating_2,
                'rating_3': summary.rating_3,
                'rating_4': summary.rating_4,
                'rating_5': summary.rating_5,
            }

        context.update({
            'review_summary': default_summary,  # Используем словарь вместо объекта
            'recent_reviews': recent_reviews,
            'attention_reviews': attention_reviews,
            'status_stats': status_stats,
            'seller': seller,
        })
        return context

    def get_status_stats(self, seller):
        stats = ProductReview.objects.filter(
            product__seller=seller
        ).values('status').annotate(count=Count('id'))

        total = sum(stat['count'] for stat in stats)

        for stat in stats:
            stat['percentage'] = (stat['count'] / total * 100) if total > 0 else 0

        return stats


class ReviewsListView(LoginRequiredMixin, ListView):
    template_name = 'sellers/reviews/list.html'
    context_object_name = 'reviews'
    paginate_by = 20

    def get_queryset(self):
        seller = self.request.user.seller
        queryset = ProductReview.objects.filter(product__seller=seller)

        # Фильтрация
        rating = self.request.GET.get('rating')
        status = self.request.GET.get('status')
        has_response = self.request.GET.get('has_response')
        search = self.request.GET.get('search')

        if rating:
            queryset = queryset.filter(rating=rating)
        if status:
            queryset = queryset.filter(status=status)
        if has_response:
            if has_response == 'yes':
                queryset = queryset.exclude(seller_response='')
            else:
                queryset = queryset.filter(seller_response='')
        if search:
            queryset = queryset.filter(
                Q(product__name__icontains=search) |
                Q(comment__icontains=search) |
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search)
            )

        return queryset.select_related('product', 'user').prefetch_related('media_files')  # ИЗМЕНЕНО

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        seller = self.request.user.seller

        context['total_count'] = self.get_queryset().count()
        context['rating_stats'] = self.get_rating_stats(seller)
        context['seller'] = seller

        return context

    def get_rating_stats(self, seller):
        return ProductReview.objects.filter(
            product__seller=seller,
            status='approved'
        ).values('rating').annotate(count=Count('id')).order_by('-rating')


class ReviewResponseView(LoginRequiredMixin, View):
    def post(self, request, review_id):
        seller = request.user.seller
        review = get_object_or_404(ProductReview, id=review_id, product__seller=seller)

        data = json.loads(request.body)
        response_text = data.get('response', '').strip()

        if not response_text:
            return JsonResponse({
                'success': False,
                'error': 'Текст ответа не может быть пустым'
            })

        # Обновляем отзыв
        review.seller_response = response_text
        review.responded_at = timezone.now()
        review.status = 'edited'
        review.save()

        # Обновляем статистику
        summary = SellerReviewSummary.objects.get(seller=seller)
        summary.update_statistics()

        return JsonResponse({
            'success': True,
            'message': 'Ответ успешно добавлен',
            'response': review.seller_response,
            'responded_at': review.responded_at.strftime('%d.%m.%Y %H:%M')
        })


class ReviewReportView(LoginRequiredMixin, View):
    def post(self, request, review_id):
        seller = request.user.seller
        review = get_object_or_404(ProductReview, id=review_id, product__seller=seller)

        data = json.loads(request.body)
        report_type = data.get('report_type')
        description = data.get('description', '')

        # Создаем жалобу
        report = ReviewReport.objects.create(
            review=review,
            reported_by=request.user,
            report_type=report_type,
            description=description
        )

        # Меняем статус отзыва на модерацию
        review.status = 'pending'
        review.save()

        return JsonResponse({
            'success': True,
            'message': 'Жалоба отправлена на модерацию'
        })


class ReviewAnalyticsView(LoginRequiredMixin, TemplateView):
    template_name = 'sellers/reviews/analytics.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        seller = self.request.user.seller

        # Общая статистика
        summary = SellerReviewSummary.objects.get(seller=seller)

        # Динамика отзывов за последние 6 месяцев
        monthly_stats = self.get_monthly_stats(seller)

        # Рейтинг товаров
        product_ratings = self.get_product_ratings(seller)

        context.update({
            'review_summary': summary,
            'monthly_stats': monthly_stats,
            'product_ratings': product_ratings,
            'seller': seller,
        })
        return context

    def get_monthly_stats(self, seller):
        from django.db.models.functions import TruncMonth
        from django.db.models import Count

        return ProductReview.objects.filter(
            product__seller=seller,
            status='approved',
            created_at__gte=timezone.now() - timedelta(days=180)
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            count=Count('id'),
            avg_rating=Avg('rating')
        ).order_by('month')

    def get_product_ratings(self, seller):
        return ProductReview.objects.filter(
            product__seller=seller,
            status='approved'
        ).values('product__name').annotate(
            avg_rating=Avg('rating'),
            review_count=Count('id')
        ).order_by('-avg_rating')[:10]


class ReviewsAnalyticsView(LoginRequiredMixin, TemplateView):
    template_name = 'sellers/reviews/analytics.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        seller = self.request.user.seller

        # Получаем сводную статистику
        summary, created = SellerReviewSummary.objects.get_or_create(seller=seller)
        if created:
            summary.update_statistics()

        # Динамика отзывов за последние 6 месяцев
        monthly_stats = self.get_monthly_stats(seller)

        # Динамика за последние 12 недель
        weekly_stats = self.get_weekly_stats(seller)

        # Рейтинг товаров
        product_ratings = self.get_product_ratings(seller)

        # Анализ ответов
        response_analytics = self.get_response_analytics(seller)

        # Распределение по рейтингам
        rating_distribution = self.get_rating_distribution(seller)

        # Рекомендации
        recommendations = self.get_recommendations(seller, summary)

        context.update({
            'review_summary': summary,
            'monthly_stats': monthly_stats,
            'weekly_stats': weekly_stats,
            'product_ratings': product_ratings,
            'response_analytics': response_analytics,
            'rating_distribution': rating_distribution,
            'recommendations': recommendations,
            'seller': seller,
        })
        return context

    def get_monthly_stats(self, seller):
        """Статистика отзывов по месяцам за последние 6 месяцев"""
        six_months_ago = timezone.now() - timedelta(days=180)

        monthly_data = ProductReview.objects.filter(
            product__seller=seller,
            status='approved',
            created_at__gte=six_months_ago
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total_reviews=Count('id'),
            avg_rating=Avg('rating'),
            positive_reviews=Count('id', filter=Q(rating__gte=4)),
            negative_reviews=Count('id', filter=Q(rating__lte=2))
        ).order_by('month')

        # Форматируем данные для графика
        months = []
        totals = []
        ratings = []
        positive_rates = []

        for data in monthly_data:
            months.append(data['month'].strftime('%b %Y'))
            totals.append(data['total_reviews'])
            ratings.append(float(data['avg_rating']))

            # Процент положительных отзывов
            total = data['total_reviews']
            positive = data['positive_reviews']
            positive_rate = (positive / total * 100) if total > 0 else 0
            positive_rates.append(round(positive_rate, 1))

        return {
            'labels': months,
            'datasets': [
                {
                    'label': 'Всего отзывов',
                    'data': totals,
                    'borderColor': '#667eea',
                    'backgroundColor': 'rgba(102, 126, 234, 0.1)',
                },
                {
                    'label': 'Средний рейтинг',
                    'data': ratings,
                    'borderColor': '#10b981',
                    'backgroundColor': 'transparent',
                    'yAxisID': 'y1'
                },
                {
                    'label': 'Положительные отзывы %',
                    'data': positive_rates,
                    'borderColor': '#f59e0b',
                    'backgroundColor': 'transparent',
                    'yAxisID': 'y2'
                }
            ]
        }

    def get_weekly_stats(self, seller):
        """Статистика отзывов по неделям за последние 12 недель"""
        twelve_weeks_ago = timezone.now() - timedelta(weeks=12)

        weekly_data = ProductReview.objects.filter(
            product__seller=seller,
            status='approved',
            created_at__gte=twelve_weeks_ago
        ).annotate(
            week=TruncWeek('created_at')
        ).values('week').annotate(
            total_reviews=Count('id'),
            avg_rating=Avg('rating')
        ).order_by('week')

        weeks = []
        totals = []

        for data in weekly_data:
            weeks.append(data['week'].strftime('%d.%m'))
            totals.append(data['total_reviews'])

        return {
            'labels': weeks,
            'totals': totals
        }

    def get_product_ratings(self, seller, limit=10):
        """Рейтинг товаров по отзывам"""
        return ProductReview.objects.filter(
            product__seller=seller,
            status='approved'
        ).values(
            'product__id',
            'product__name',
            'product__image'
        ).annotate(
            avg_rating=Avg('rating'),
            total_reviews=Count('id'),
            positive_reviews=Count('id', filter=Q(rating__gte=4)),
            negative_reviews=Count('id', filter=Q(rating__lte=2))
        ).filter(total_reviews__gte=1).order_by('-avg_rating')[:limit]

    def get_response_analytics(self, seller):
        """Аналитика ответов на отзывы"""
        total_reviews = ProductReview.objects.filter(
            product__seller=seller,
            status='approved'
        ).count()

        responded_reviews = ProductReview.objects.filter(
            product__seller=seller,
            status='approved'
        ).exclude(seller_response='').count()

        # Время ответа
        response_times = ProductReview.objects.filter(
            product__seller=seller,
            status='approved',
            responded_at__isnull=False
        ).annotate(
            response_days=(F('responded_at') - F('created_at'))
        ).aggregate(
            avg_days=Avg('response_days')
        )

        avg_response_days = response_times['avg_days'].days if response_times['avg_days'] else 0

        # Распределение по времени ответа
        response_time_distribution = {
            'within_day': ProductReview.objects.filter(
                product__seller=seller,
                status='approved',
                responded_at__isnull=False,
                responded_at__lte=F('created_at') + timedelta(days=1)
            ).count(),
            'within_week': ProductReview.objects.filter(
                product__seller=seller,
                status='approved',
                responded_at__isnull=False,
                responded_at__lte=F('created_at') + timedelta(days=7),
                responded_at__gt=F('created_at') + timedelta(days=1)
            ).count(),
            'over_week': ProductReview.objects.filter(
                product__seller=seller,
                status='approved',
                responded_at__isnull=False,
                responded_at__gt=F('created_at') + timedelta(days=7)
            ).count()
        }

        return {
            'total_reviews': total_reviews,
            'responded_reviews': responded_reviews,
            'response_rate': (responded_reviews / total_reviews * 100) if total_reviews > 0 else 0,
            'avg_response_days': avg_response_days,
            'response_time_distribution': response_time_distribution
        }

    def get_rating_distribution(self, seller):
        """Распределение отзывов по рейтингам"""
        distribution = ProductReview.objects.filter(
            product__seller=seller,
            status='approved'
        ).values('rating').annotate(
            count=Count('id')
        ).order_by('-rating')

        # Создаем полное распределение от 1 до 5
        full_distribution = {}
        for rating in range(1, 6):
            full_distribution[rating] = 0

        for item in distribution:
            full_distribution[item['rating']] = item['count']

        # Рассчитываем проценты
        total = sum(full_distribution.values())
        percentages = {}
        for rating, count in full_distribution.items():
            percentages[rating] = (count / total * 100) if total > 0 else 0

        return {
            'counts': full_distribution,
            'percentages': percentages,
            'total': total
        }

    def get_recommendations(self, seller, summary):
        """Рекомендации по улучшению отзывов"""
        recommendations = []

        # Анализ процента ответов
        if summary.response_rate < 80:
            recommendations.append({
                'type': 'response_rate',
                'title': 'Увеличьте процент ответов на отзывы',
                'description': f'Текущий процент ответов: {summary.response_rate:.1f}%. Старайтесь отвечать на 80%+ отзывов.',
                'priority': 'high',
                'icon': 'bi-chat-dots',
                'action': 'Ответить на отзывы'
            })

        # Анализ среднего рейтинга
        if summary.average_rating < 4.0:
            recommendations.append({
                'type': 'low_rating',
                'title': 'Улучшите средний рейтинг',
                'description': f'Текущий средний рейтинг: {summary.average_rating:.1f}. Проанализируйте негативные отзывы.',
                'priority': 'high',
                'icon': 'bi-star',
                'action': 'Анализ отзывов'
            })

        # Анализ времени ответа
        response_analytics = self.get_response_analytics(seller)
        if response_analytics['avg_response_days'] > 3:
            recommendations.append({
                'type': 'slow_response',
                'title': 'Ускорьте ответы на отзывы',
                'description': f'Среднее время ответа: {response_analytics["avg_response_days"]} дней. Старайтесь отвечать в течение 1-2 дней.',
                'priority': 'medium',
                'icon': 'bi-clock',
                'action': 'Ускорить ответы'
            })

        # Анализ негативных отзывов
        negative_reviews = ProductReview.objects.filter(
            product__seller=seller,
            rating__lte=2,
            status='approved'
        ).count()

        if negative_reviews >= 5:
            recommendations.append({
                'type': 'negative_reviews',
                'title': 'Обратите внимание на негативные отзывы',
                'description': f'Количество негативных отзывов: {negative_reviews}. Проанализируйте причины и улучшите сервис.',
                'priority': 'high',
                'icon': 'bi-exclamation-triangle',
                'action': 'Анализ проблем'
            })

        # Анализ товаров с низким рейтингом
        low_rated_products = self.get_product_ratings(seller, limit=5)
        low_rated_products = [p for p in low_rated_products if p['avg_rating'] < 3.5]

        if low_rated_products:
            recommendations.append({
                'type': 'low_rated_products',
                'title': 'Товары с низким рейтингом',
                'description': f'Найдено {len(low_rated_products)} товаров с рейтингом ниже 3.5. Улучшите их качество или описание.',
                'priority': 'medium',
                'icon': 'bi-box',
                'action': 'Просмотреть товары'
            })

        return recommendations


def get_rating_change(seller, current_rating):
    """Расчет изменения рейтинга за последний месяц"""
    from datetime import timedelta
    from django.utils import timezone

    today = timezone.now().date()
    month_ago = today - timedelta(days=30)
    two_months_ago = month_ago - timedelta(days=30)

    try:
        # Средний рейтинг за предыдущий месяц
        previous_month_reviews = ProductReview.objects.filter(
            product__seller=seller,
            status='approved',
            created_at__date__gte=two_months_ago,
            created_at__date__lt=month_ago
        )

        if previous_month_reviews.exists():
            previous_rating = previous_month_reviews.aggregate(
                avg_rating=Avg('rating')
            )['avg_rating']

            if previous_rating:
                return round(float(current_rating) - float(previous_rating), 1)

        return 0.2  # Дефолтное положительное изменение если нет данных

    except Exception as e:
        print(f"Error calculating rating change: {e}")
        return 0.2


##### Настройки ###########


class SellerSettingsView(LoginRequiredMixin, TemplateView):
    template_name = 'sellers/settings/settings.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        seller = self.request.user.seller

        # Получаем или создаем настройки
        settings, created = SellerSettings.objects.get_or_create(seller=seller)
        integrations, integrations_created = SellerIntegration.objects.get_or_create(seller=seller)

        # Непрочитанные уведомления
        unread_notifications = SellerNotification.objects.filter(
            seller=seller,
            is_read=False
        )[:5]

        # Валидация активной вкладки
        valid_tabs = ['general', 'notifications', 'automation', 'integrations', 'security', 'notifications_list']
        active_tab = self.request.GET.get('tab', 'general')
        if active_tab not in valid_tabs:
            active_tab = 'general'

        context.update({
            'seller': seller,
            'settings': settings,
            'integrations': integrations,
            'unread_notifications': unread_notifications,
            'active_tab': active_tab
        })
        return context


class UpdateSettingsView(LoginRequiredMixin, View):
    def post(self, request):
        seller = request.user.seller
        settings, created = SellerSettings.objects.get_or_create(seller=seller)

        try:
            data = json.loads(request.body)
            setting_type = data.get('type')

            if setting_type == 'general':
                # Общие настройки
                settings.email_notifications = data.get('email_notifications', True)
                settings.sms_notifications = data.get('sms_notifications', False)
                settings.push_notifications = data.get('push_notifications', True)
                settings.theme = data.get('theme', 'light')
                settings.language = data.get('language', 'ru')

            elif setting_type == 'notifications':
                # Уведомления
                settings.notify_new_orders = data.get('notify_new_orders', True)
                settings.notify_low_stock = data.get('notify_low_stock', True)
                settings.notify_new_reviews = data.get('notify_new_reviews', True)
                settings.notify_payouts = data.get('notify_payouts', True)

            elif setting_type == 'automation':
                # Автоматизация
                settings.auto_confirm_orders = data.get('auto_confirm_orders', False)
                settings.auto_update_stock = data.get('auto_update_stock', False)
                settings.auto_response_reviews = data.get('auto_response_reviews', False)

            elif setting_type == 'security':
                # Безопасность
                settings.two_factor_auth = data.get('two_factor_auth', False)
                settings.login_alerts = data.get('login_alerts', True)

            settings.save()

            return JsonResponse({
                'success': True,
                'message': 'Настройки успешно обновлены'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })


class UpdateIntegrationsView(LoginRequiredMixin, View):
    def post(self, request):
        seller = request.user.seller
        integrations, created = SellerIntegration.objects.get_or_create(seller=seller)

        try:
            data = json.loads(request.body)

            # Google Analytics
            integrations.google_analytics = data.get('google_analytics', False)
            integrations.google_analytics_id = data.get('google_analytics_id', '')

            # Яндекс.Метрика
            integrations.yandex_metrika = data.get('yandex_metrika', False)
            integrations.yandex_metrika_id = data.get('yandex_metrika_id', '')

            # Экспорт
            integrations.auto_export_orders = data.get('auto_export_orders', False)
            integrations.export_format = data.get('export_format', 'csv')

            integrations.save()

            return JsonResponse({
                'success': True,
                'message': 'Настройки интеграций обновлены'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })


class GenerateApiKeyView(LoginRequiredMixin, View):
    def post(self, request):
        seller = request.user.seller
        integrations, created = SellerIntegration.objects.get_or_create(seller=seller)

        try:
            import secrets
            import string

            # Генерация API ключа
            alphabet = string.ascii_letters + string.digits
            api_key = 'avl_' + ''.join(secrets.choice(alphabet) for i in range(32))
            api_secret = 'avls_' + ''.join(secrets.choice(alphabet) for i in range(48))

            integrations.api_key = api_key
            integrations.api_secret = api_secret
            integrations.save()

            return JsonResponse({
                'success': True,
                'api_key': api_key,
                'api_secret': api_secret,
                'message': 'Новые API ключи сгенерированы'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })


class MarkNotificationReadView(LoginRequiredMixin, View):
    def post(self, request, notification_id):
        seller = request.user.seller

        try:
            notification = SellerNotification.objects.get(
                id=notification_id,
                seller=seller
            )
            notification.is_read = True
            notification.save()

            return JsonResponse({
                'success': True,
                'message': 'Уведомление помечено как прочитанное'
            })

        except SellerNotification.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Уведомление не найдено'
            })


class ClearAllNotificationsView(LoginRequiredMixin, View):
    def post(self, request):
        seller = request.user.seller

        try:
            SellerNotification.objects.filter(seller=seller).update(is_read=True)

            return JsonResponse({
                'success': True,
                'message': 'Все уведомления помечены как прочитанные'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })


# Профиль

class SellerProfileView(LoginRequiredMixin, TemplateView):
    template_name = 'sellers/seller_profile.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        seller = self.request.user.seller

        # Получаем непрочитанные уведомления (если есть такая модель)
        try:
            from .models import SellerNotification
            unread_notifications = SellerNotification.objects.filter(
                seller=seller,
                is_read=False
            )[:5]
        except:
            unread_notifications = []

        context.update({
            'seller': seller,
            'unread_notifications': unread_notifications
        })
        return context


class UpdateContactDataView(LoginRequiredMixin, View):
    def post(self, request):
        seller = request.user.seller
        user = request.user

        try:
            # Валидация email
            new_email = request.POST.get('email', '').strip()
            if new_email and new_email != user.email:
                if User.objects.filter(email=new_email).exclude(id=user.id).exists():
                    return JsonResponse({
                        'success': False,
                        'error': 'Этот email уже используется другим пользователем'
                    })

            # Валидация телефона
            new_phone = request.POST.get('phone', '').strip()
            if new_phone and new_phone != seller.phone:
                # Проверяем, не используется ли телефон другим продавцом
                from .models import Seller
                if Seller.objects.filter(phone=new_phone).exclude(id=seller.id).exists():
                    return JsonResponse({
                        'success': False,
                        'error': 'Этот телефон уже используется другим продавцом'
                    })

            # Обновляем данные продавца
            seller.contact_person = request.POST.get('contact_person', '').strip()
            seller.phone = new_phone
            seller.save()

            # Обновляем email пользователя
            if new_email:
                user.email = new_email
                user.save()

            return JsonResponse({
                'success': True,
                'message': 'Контактные данные успешно обновлены'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Ошибка при сохранении: {str(e)}'
            })


# API endpoints для добавления товара

@login_required
def get_category_children(request, category_id):
    """Получить дочерние категории"""
    try:
        category = Category.objects.get(id=category_id, is_active=True)
        children = Category.objects.filter(parent=category, is_active=True).order_by('order', 'name')
        return JsonResponse([
            {'id': cat.id, 'name': cat.name}
            for cat in children
        ], safe=False)
    except Category.DoesNotExist:
        return JsonResponse([], safe=False)


@login_required
def get_vin_info(request):
    """
    Получить информацию о транспорте по VIN
    НОВАЯ ВЕРСИЯ: использует оптимизированный VIN декодер с кэшем
    """
    vin = VINService.normalize_vin(request.GET.get('vin', ''))
    vin_validation = VINService.validate_vin(vin) if vin else {
        'valid': False,
        'error': 'VIN обязателен',
        'normalized_vin': vin,
    }

    if not vin_validation['valid']:
        return JsonResponse({
            'success': False,
            'error': vin_validation['error'],
            'normalized_vin': vin_validation.get('normalized_vin') or vin,
        })

    # Декодировать VIN (с кэшем!)
    try:
        from vehicles.vin_decoder import VINDecoder
        decoded = VINDecoder.decode(vin)
    except Exception:
        decoded = None
    
    if not decoded:
        # Fallback на старый метод (если VIN не декодирован)
        return _get_vin_info_old(request, vin)
    
    best = decoded['best_match']
    
    if not best:
        return JsonResponse({
            'success': False,
            'error': 'Модификация не определена'
        })
    
    # Формировать информацию о ТС
    vehicle_info = []
    vehicle_data = {
        'vehicle_type': best.vehicle_type.code,
        'manufacturer_name': decoded['manufacturer'].name,
        'modification_id': best.id,
        'modification_name': best.name,
        'year': decoded['year'],
    }
    
    vehicle_info.append(f"Производитель: {decoded['manufacturer'].name}")
    vehicle_info.append(f"Модификация: {best.name}")
    
    if decoded['year']:
        vehicle_info.append(f"Год: {decoded['year']}")
    
    if best.engine_volume:
        vehicle_info.append(f"Объем: {best.engine_volume} см³")
    
    if best.power:
        vehicle_info.append(f"Мощность: {best.power} л.с.")
    
    return JsonResponse({
        'success': True,
        'normalized_vin': vin,
        'vehicle_info': '<br>'.join(vehicle_info),
        'vehicle_data': vehicle_data
    })


def _get_vin_info_old(request, vin):
    """
    СТАРЫЙ метод получения информации по VIN (fallback)
    Используется если новый декодер не смог распознать VIN
    """
    from shop.models import (
        Product, CarBrand, TruckBrand, MotoBrand, SpecialBrand,
        CarModel, TruckModel, MotoModel, SpecialModel,
        CarGeneration, TruckGeneration, MotoGeneration, SpecialGeneration,
        CarModification, TruckModification, MotoModification, SpecialModification
    )
    
    products = Product.objects.filter(vin=vin).select_related(
        'car_brand', 'truck_brand', 'moto_brand', 'special_brand', 'category'
    ).prefetch_related(
        'car_models', 'car_generations', 'car_modifications',
        'truck_models', 'truck_generations', 'truck_modifications',
        'moto_models', 'moto_generations', 'moto_modifications',
        'special_models', 'special_generations', 'special_modifications'
    )[:1]
    
    if products.exists():
        product = products.first()
        vehicle_info = []
        vehicle_data = {
            'vehicle_type': product.vehicle_type if product.vehicle_type != 'universal' else None,
            'brand_id': None,
            'brand_name': None,
            'model_id': None,
            'model_name': None,
            'generation_id': None,
            'generation_name': None,
            'modification_id': None,
            'modification_name': None,
        }
        
        # Определяем тип транспорта и заполняем данные
        if product.vehicle_type == 'car' and product.car_brand:
            vehicle_data['brand_id'] = product.car_brand.id
            vehicle_data['brand_name'] = product.car_brand.name
            vehicle_info.append(f"Марка: {product.car_brand.name}")
            
            if product.car_models.exists():
                model = product.car_models.first()
                vehicle_data['model_id'] = model.id
                vehicle_data['model_name'] = model.name
                vehicle_info.append(f"Модель: {model.name}")
            
            if product.car_generations.exists():
                generation = product.car_generations.first()
                vehicle_data['generation_id'] = generation.id
                vehicle_data['generation_name'] = generation.name
                vehicle_info.append(f"Поколение: {generation.name}")
            
            if product.car_modifications.exists():
                modification = product.car_modifications.first()
                vehicle_data['modification_id'] = modification.id
                vehicle_data['modification_name'] = modification.name
                vehicle_info.append(f"Модификация: {modification.name}")
        
        elif product.vehicle_type == 'truck' and product.truck_brand:
            vehicle_data['brand_id'] = product.truck_brand.id
            vehicle_data['brand_name'] = product.truck_brand.name
            vehicle_info.append(f"Марка: {product.truck_brand.name}")
            
            if product.truck_models.exists():
                model = product.truck_models.first()
                vehicle_data['model_id'] = model.id
                vehicle_data['model_name'] = model.name
                vehicle_info.append(f"Модель: {model.name}")
            
            if product.truck_generations.exists():
                generation = product.truck_generations.first()
                vehicle_data['generation_id'] = generation.id
                vehicle_data['generation_name'] = generation.name
                vehicle_info.append(f"Поколение: {generation.name}")
            
            if product.truck_modifications.exists():
                modification = product.truck_modifications.first()
                vehicle_data['modification_id'] = modification.id
                vehicle_data['modification_name'] = modification.name
                vehicle_info.append(f"Модификация: {modification.name}")
        
        elif product.vehicle_type == 'moto' and product.moto_brand:
            vehicle_data['brand_id'] = product.moto_brand.id
            vehicle_data['brand_name'] = product.moto_brand.name
            vehicle_info.append(f"Марка: {product.moto_brand.name}")
            
            if product.moto_models.exists():
                model = product.moto_models.first()
                vehicle_data['model_id'] = model.id
                vehicle_data['model_name'] = model.name
                vehicle_info.append(f"Модель: {model.name}")
            
            if product.moto_generations.exists():
                generation = product.moto_generations.first()
                vehicle_data['generation_id'] = generation.id
                vehicle_data['generation_name'] = generation.name
                vehicle_info.append(f"Поколение: {generation.name}")
            
            if product.moto_modifications.exists():
                modification = product.moto_modifications.first()
                vehicle_data['modification_id'] = modification.id
                vehicle_data['modification_name'] = modification.name
                vehicle_info.append(f"Модификация: {modification.name}")
        
        elif product.vehicle_type == 'special' and product.special_brand:
            vehicle_data['brand_id'] = product.special_brand.id
            vehicle_data['brand_name'] = product.special_brand.name
            vehicle_info.append(f"Марка: {product.special_brand.name}")
            
            if product.special_models.exists():
                model = product.special_models.first()
                vehicle_data['model_id'] = model.id
                vehicle_data['model_name'] = model.name
                vehicle_info.append(f"Модель: {model.name}")
            
            if product.special_generations.exists():
                generation = product.special_generations.first()
                vehicle_data['generation_id'] = generation.id
                vehicle_data['generation_name'] = generation.name
                vehicle_info.append(f"Поколение: {generation.name}")
            
            if product.special_modifications.exists():
                modification = product.special_modifications.first()
                vehicle_data['modification_id'] = modification.id
                vehicle_data['modification_name'] = modification.name
                vehicle_info.append(f"Модификация: {modification.name}")
        
        if product.category:
            vehicle_info.append(f"Категория: {product.category.name}")
        
        return JsonResponse({
            'success': True,
            'vehicle_info': '<br>'.join(vehicle_info) if vehicle_info else 'Информация найдена, но детали недоступны',
            'vehicle_data': vehicle_data
        })
    
    # Если не найдено, можно попробовать декодировать VIN (упрощенная версия)
    # WMI (первые 3 символа) могут указывать на производителя
    wmi = vin[:3]
    wmi_to_brand = {
        'WAU': 'Audi', 'TRU': 'Audi',
        'WBA': 'BMW', 'WBS': 'BMW', 'WBX': 'BMW',
        'WDB': 'Mercedes-Benz', 'WDD': 'Mercedes-Benz', 'WDF': 'Mercedes-Benz',
        'WVW': 'Volkswagen', 'WVG': 'Volkswagen', '3VW': 'Volkswagen',
        'JTD': 'Toyota', 'JTE': 'Toyota', '5TE': 'Toyota',
        'JHM': 'Honda', '1HG': 'Honda', '2HG': 'Honda',
        '1FA': 'Ford', '1FB': 'Ford', '1FT': 'Ford',
        'KMH': 'Hyundai', '5NP': 'Hyundai',
    }
    
    brand_name = wmi_to_brand.get(wmi)
    if brand_name:
        return JsonResponse({
            'success': True,
            'vehicle_info': f'Возможная марка: {brand_name} (определено по WMI)',
            'vehicle_data': None
        })
    
    return JsonResponse({
        'success': False,
        'error': 'VIN не найден в базе'
    })


@login_required
def get_vehicle_brands(request):
    """Получить список марок транспорта"""
    vehicle_type = request.GET.get('type')
    
    if vehicle_type == 'car':
        from shop.models import CarBrand
        brands = CarBrand.objects.filter(is_active=True).order_by('name')
        return JsonResponse([
            {'id': brand.id, 'name': brand.name}
            for brand in brands
        ], safe=False)
    elif vehicle_type == 'truck':
        from shop.models import TruckBrand
        brands = TruckBrand.objects.filter(is_active=True).order_by('name')
        return JsonResponse([
            {'id': brand.id, 'name': brand.name}
            for brand in brands
        ], safe=False)
    elif vehicle_type == 'moto':
        from shop.models import MotoBrand
        brands = MotoBrand.objects.filter(is_active=True).order_by('name')
        return JsonResponse([
            {'id': brand.id, 'name': brand.name}
            for brand in brands
        ], safe=False)
    elif vehicle_type == 'special':
        from shop.models import SpecialBrand
        brands = SpecialBrand.objects.filter(is_active=True).order_by('name')
        return JsonResponse([
            {'id': brand.id, 'name': brand.name}
            for brand in brands
        ], safe=False)
    
    return JsonResponse([], safe=False)


@login_required
def get_vehicle_models(request):
    """Получить список моделей транспорта"""
    vehicle_type = request.GET.get('type')
    brand_id = request.GET.get('brand')
    
    if not brand_id:
        return JsonResponse([], safe=False)
    
    if vehicle_type == 'car':
        from shop.models import CarModel
        models = CarModel.objects.filter(brand_id=brand_id, is_active=True).order_by('name')
        return JsonResponse([
            {'id': model.id, 'name': model.name}
            for model in models
        ], safe=False)
    elif vehicle_type == 'truck':
        from shop.models import TruckModel
        models = TruckModel.objects.filter(brand_id=brand_id, is_active=True).order_by('name')
        return JsonResponse([
            {'id': model.id, 'name': model.name}
            for model in models
        ], safe=False)
    elif vehicle_type == 'moto':
        from shop.models import MotoModel
        models = MotoModel.objects.filter(brand_id=brand_id, is_active=True).order_by('name')
        return JsonResponse([
            {'id': model.id, 'name': model.name}
            for model in models
        ], safe=False)
    elif vehicle_type == 'special':
        from shop.models import SpecialModel
        models = SpecialModel.objects.filter(brand_id=brand_id, is_active=True).order_by('name')
        return JsonResponse([
            {'id': model.id, 'name': model.name}
            for model in models
        ], safe=False)
    
    return JsonResponse([], safe=False)


@login_required
def get_vehicle_generations(request):
    """Получить список поколений транспорта"""
    vehicle_type = request.GET.get('type')
    model_id = request.GET.get('model')
    
    if not model_id:
        return JsonResponse([], safe=False)
    
    if vehicle_type == 'car':
        from shop.models import CarGeneration
        generations = CarGeneration.objects.filter(model_id=model_id, is_active=True).order_by('-year_start')
        return JsonResponse([
            {'id': gen.id, 'name': gen.name}
            for gen in generations
        ], safe=False)
    elif vehicle_type == 'truck':
        from shop.models import TruckGeneration
        generations = TruckGeneration.objects.filter(model_id=model_id, is_active=True).order_by('-year_start')
        return JsonResponse([
            {'id': gen.id, 'name': gen.name}
            for gen in generations
        ], safe=False)
    elif vehicle_type == 'moto':
        from shop.models import MotoGeneration
        generations = MotoGeneration.objects.filter(model_id=model_id, is_active=True).order_by('-year_start')
        return JsonResponse([
            {'id': gen.id, 'name': gen.name}
            for gen in generations
        ], safe=False)
    elif vehicle_type == 'special':
        from shop.models import SpecialGeneration
        generations = SpecialGeneration.objects.filter(model_id=model_id, is_active=True).order_by('-year_start')
        return JsonResponse([
            {'id': gen.id, 'name': gen.name}
            for gen in generations
        ], safe=False)
    
    return JsonResponse([], safe=False)


@login_required
def get_vehicle_modifications(request):
    """Получить список модификаций транспорта"""
    vehicle_type = request.GET.get('type')
    generation_id = request.GET.get('generation')
    
    if not generation_id:
        return JsonResponse([], safe=False)
    
    if vehicle_type == 'car':
        from shop.models import CarModification
        modifications = CarModification.objects.filter(generation_id=generation_id).order_by('name')
        return JsonResponse([
            {'id': mod.id, 'name': mod.name}
            for mod in modifications
        ], safe=False)
    elif vehicle_type == 'truck':
        from shop.models import TruckModification
        modifications = TruckModification.objects.filter(generation_id=generation_id).order_by('name')
        return JsonResponse([
            {'id': mod.id, 'name': mod.name}
            for mod in modifications
        ], safe=False)
    elif vehicle_type == 'moto':
        from shop.models import MotoModification
        modifications = MotoModification.objects.filter(generation_id=generation_id).order_by('name')
        return JsonResponse([
            {'id': mod.id, 'name': mod.name}
            for mod in modifications
        ], safe=False)
    elif vehicle_type == 'special':
        from shop.models import SpecialModification
        modifications = SpecialModification.objects.filter(generation_id=generation_id).order_by('name')
        return JsonResponse([
            {'id': mod.id, 'name': mod.name}
            for mod in modifications
        ], safe=False)

    return JsonResponse([], safe=False)


@login_required
def suggest_category(request):
    """Продавец предлагает новую категорию/подкатегорию для модерации."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не разрешён'}, status=405)

    try:
        seller = request.user.seller
    except Seller.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Нет прав продавца'})

    name = request.POST.get('name', '').strip()
    if not name:
        return JsonResponse({'success': False, 'error': 'Укажите название категории'})

    if len(name) > 255:
        return JsonResponse({'success': False, 'error': 'Название слишком длинное (максимум 255 символов)'})

    parent_id = request.POST.get('parent_category_id') or None
    parent_category = None
    if parent_id:
        try:
            parent_category = Category.objects.get(id=parent_id, is_active=True)
        except Category.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Родительская категория не найдена'})

    description = request.POST.get('description', '').strip()

    duplicate = CategoryRequest.objects.filter(
        seller=seller, name__iexact=name, parent_category=parent_category, status='pending'
    ).exists()
    if duplicate:
        return JsonResponse({'success': False, 'error': 'Такая заявка уже отправлена и ожидает проверки'})

    CategoryRequest.objects.create(
        seller=seller,
        name=name,
        parent_category=parent_category,
        description=description,
    )
    return JsonResponse({'success': True, 'message': 'Заявка отправлена на модерацию. Мы рассмотрим её в ближайшее время.'})
